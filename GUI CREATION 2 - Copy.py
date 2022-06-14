##### TO DO
#####  CREATE EXCUTE FUNCTION FOR SIGNING OFF PASSEDINVOCIES ONTO EXCEL SHEET
##### ALLOW OPTION TO MNAUALLY SAERCH AND ASSIGN PO NUMBERS TO INVOCIES ALREADY CHECKED


import collections, pytesseract, os, PyPDF3, openpyxl, re, sys, shutil
import tkinter
from tkinter import *
from tkinter import filedialog, filedialog
from tkPDFViewer import tkPDFViewer as pdf
from openpyxl.utils.cell import get_column_letter
from pdf2image import convert_from_path
from PIL import Image
from datetime import date
pytesseract.pytesseract.tesseract_cmd = r'C:\\Users\\nickb\\Desktop\\Python\\Tesseract-OCR\\tesseract.exe'
#poppler_path = r'C:\\Users\\nbolade\\Desktop\\Main Folder\\Core Files\\poppler-0.68.0_x86\\poppler-0.68.0\\bin'  ###change this manually in both locations
poppler_path = r'C:\\Users\\nickb\\Desktop\\Python\\poppler-0.68.0_x86\\poppler-0.68.0\\bin'
root=Tk()
root.title("THE ORGINAL FACTORY SHOP INVOICE CHECKER")
root.geometry("1920x1000+0+0")
root.configure(bg="grey")
today = date.today()
DateToday = today.strftime("%d.%m.%y")
DateToday2 = today.strftime("%d/%m/%y")
#root.tk.call('tk', 'scaling', 1)
try:
    os.chdir('C:\\Users\\nbolade\\Desktop\\Main Folder\\Invoice checking folder')
    InvoicesInFolderList = []
    FileFolderAllList = os.listdir("C:\\Users\\nbolade\\Desktop\\Main Folder\\Invoice checking folder")
    FileFolderToCheck = "C:\\Users\\nbolade\\Desktop\\Main Folder\\Invoice checking folder"
    pdffilepath = "C:\\Users\\nbolade\\Desktop\\Main Folder\\Invoice checking folder\\"
    MainFolder = "C:\\Users\\nbolade\\Desktop\\Main Folder\\"
    ToCheckFolder = "C:\\Users\\nbolade\\Desktop\\Main Folder\\To Check\\"
    PassedFolder = "C:\\Users\\nbolade\\Desktop\\Main Folder\\Passed\\"
    IssuesFolder = "C:\\Users\\nbolade\\Desktop\\Main Folder\\Issues\\"
    NotOnLogFolder = "C:\\Users\\nbolade\\Desktop\\Main Folder\\Not on log\\"
    for files in FileFolderAllList:
        if files.endswith('.pdf') or files.endswith('.PDF'):
            InvoicesInFolderList.append(files)
    totalInvoicesInFolder = len(InvoicesInFolderList)
except:
    os.chdir('C:\\Users\\nickb\\Desktop\\Python')
    InvoicesInFolderList = []
    for files in os.listdir("C:\\Users\\nickb\\Desktop\\Python"):
        if files.endswith('.pdf') or files.endswith('.PDF'):
            InvoicesInFolderList.append(files)
    totalInvoicesInFolder = len(InvoicesInFolderList)
    pass
###########################################################################################
passedcounter = 0
issuecounter = 0
Notonlogcounter = 0
InvoiceBeingProcessed = 0
ToCheckCounter = 0
###########################################################################################
tocheckinvoies = []        ######   needed to assign and move into folder? 
ListofPOstobePassed = []
global InvoiceNametoInvoicePONumbDict
InvoiceNametoInvoicePONumbDict = {}   ###Global Dictonairy for merging Invoicefile names and the POs in that file
def browser(): #alot of stuff in here cant be deleted, as it is done in the sort function
    global filename
    filename = filedialog.askopenfilename(initialdir = "C:\\Main Folder",title = "Select file",filetypes = (("Excel","*xlsx"),("ExcelCSV","*csv")))
    g = os.path.basename(filename)
    try:
        FileOpened.configure(text="File Opened: "+g)
        openpyxl.load_workbook(filename)
    except:
        FileOpened.configure(text = "ERROR: no file selected")
        return
    root.update()
def radioselect1():
    if ANCHOR not in str(IssueList):
        g = Passlist.get(ANCHOR)
        IssueList.insert(END,str(g))   #####set so that it can only be added once
    if ANCHOR in Passlist:
        Passlist.delete(END,str(Passlist.get(ANCHOR)))
    root.update()
def radioselect2():
    if ANCHOR in ListofPOstobePassed:
        ListofPOstobePassed.remove(ANCHOR)
    root.update()
def radioselect3():
    if ANCHOR in ListofPOstobePassed:
        ListofPOstobePassed.remove(ANCHOR)
    root.update()
def radioselect4():
    if ANCHOR in ListofPOstobePassed:
        ListofPOstobePassed.remove(ANCHOR)
    root.update()
def select(event):
    v1.img_object_li.clear()
    g =Passlist.get(ANCHOR)
    v2 = v1.pdf_view(frameLeftSide,pdf_location = r"C:\\Users\\nickb\\Desktop\\Python\\"+g, bar=False, width = 122, height = 60) # 100 by 60
    v2.place(x=0, y=0)
    root.update()
    global POSupListdict
    global POCostListdict
    global POtoProductNameDict
    global POtoQtyDict
    global SupCodeBox
    global POCodeBox
    global SKUNameBox
    global QtyBox
    global TotalPriceBox
    SupCodeBox.delete(0,END)
    POCodeBox.delete(0,END)
    SKUNameBox.delete(0,END)
    QtyBox.delete(0,END)
    TotalPriceBox.delete(0,END)
    for k in InvoiceNametoInvoicePONumbDict.get(g):
        productnamelistfordict = []
        productqtylistfordict = []
        prodictpricelistfordict = []
        for b in POtoProductNameDict.get(k):
            productnamelistfordict.append(b)
        for w in POtoQtyDict.get(k):
            productqtylistfordict.append(w)
        for q in POCostListdict.get(k):
            prodictpricelistfordict.append(q)
        productnametoqtydict = collections.defaultdict(list)
        productnametopricedict = collections.defaultdict(list)
        for key, value in zip(productnamelistfordict, productqtylistfordict): #MEEDS TO CHANGE
            productnametoqtydict[key].append(value)
        for key, value in zip(productnamelistfordict, prodictpricelistfordict):   #MEEDS TO CHANGE
            productnametopricedict[key].append(value)
        for y in POtoProductNameDict.get(k):
            SupCodeBox.insert(END,POSupListdict.get(str(k)))   
            POCodeBox.insert(END,k) 
            SKUNameBox.insert(END,y) 
            QtyBox.insert(END,str(productnametoqtydict.get(y))) 
            TotalPriceBox.insert(END,str(productnametopricedict.get(y))) 
        TotalPriceBox.insert(END,sum(POCostListdict.get(k)))     ######  INSERTS FINAL PRICE SUM (NEED TO ADD  SPACE TO EVEN OUT)
        SupCodeBox.insert(END," ")   
        POCodeBox.insert(END," ") 
        SKUNameBox.insert(END," ") 
        QtyBox.insert(END," ") 
    root.update()
    v1.img_object_li.clear()
def select2(event):
    v1.img_object_li.clear()
    g =IssueList.get(ANCHOR)
    v2 = v1.pdf_view(frameLeftSide,pdf_location = r"C:\\Users\\nickb\\Desktop\\Python\\"+g, bar=False, width = 122, height = 60) # 100 by 60
    v2.place(x=0, y=0)
    root.update()
    global POSupListdict
    global POCostListdict
    global POtoProductNameDict
    global POtoQtyDict
    global SupCodeBox
    global POCodeBox
    global SKUNameBox
    global QtyBox
    global TotalPriceBox
    SupCodeBox.delete(0,END)
    POCodeBox.delete(0,END)
    SKUNameBox.delete(0,END)
    QtyBox.delete(0,END)
    TotalPriceBox.delete(0,END)
    for k in InvoiceNametoInvoicePONumbDict.get(g):
        productnamelistfordict = []
        productqtylistfordict = []
        prodictpricelistfordict = []
        for b in POtoProductNameDict.get(k):
            productnamelistfordict.append(b)
        for w in POtoQtyDict.get(k):
            productqtylistfordict.append(w)
        for q in POCostListdict.get(k):
            prodictpricelistfordict.append(q)
        productnametoqtydict = collections.defaultdict(list)
        productnametopricedict = collections.defaultdict(list)
        for key, value in zip(productnamelistfordict, productqtylistfordict): #MEEDS TO CHANGE
            productnametoqtydict[key].append(value)
        for key, value in zip(productnamelistfordict, prodictpricelistfordict):   #MEEDS TO CHANGE
            productnametopricedict[key].append(value)
        for y in POtoProductNameDict.get(k):
            SupCodeBox.insert(END,POSupListdict.get(str(k)))   
            POCodeBox.insert(END,k) 
            SKUNameBox.insert(END,y) 
            QtyBox.insert(END,str(productnametoqtydict.get(y))) 
            TotalPriceBox.insert(END,str(productnametopricedict.get(y))) 
        TotalPriceBox.insert(END,sum(POCostListdict.get(k)))     ######  INSERTS FINAL PRICE SUM (NEED TO ADD  SPACE TO EVEN OUT)
        SupCodeBox.insert(END," ")   
        POCodeBox.insert(END," ") 
        SKUNameBox.insert(END," ") 
        QtyBox.insert(END," ") 
    root.update()
    v1.img_object_li.clear()
def select3(event):
    global v1
    v1.img_object_li.clear()
    g =Passlist.get(ANCHOR)
    v2 = v1.pdf_view(frameLeftSide,pdf_location = r"C:\\Users\\nickb\\Desktop\\Python\\"+g, bar=False, width = 122, height = 60) # 100 by 60
    v2.place(x=0, y=0)
    root.update()
    global POSupListdict
    global POCostListdict
    global POtoProductNameDict
    global POtoQtyDict
    global SupCodeBox
    global POCodeBox
    global SKUNameBox
    global QtyBox
    global TotalPriceBox
    SupCodeBox.delete(0,END)
    POCodeBox.delete(0,END)
    SKUNameBox.delete(0,END)
    QtyBox.delete(0,END)
    TotalPriceBox.delete(0,END)
    for k in InvoiceNametoInvoicePONumbDict.get(g):
        productnamelistfordict = []
        productqtylistfordict = []
        prodictpricelistfordict = []
        for b in POtoProductNameDict.get(k):
            productnamelistfordict.append(b)
        for w in POtoQtyDict.get(k):
            productqtylistfordict.append(w)
        for q in POCostListdict.get(k):
            prodictpricelistfordict.append(q)
        productnametoqtydict = collections.defaultdict(list)
        productnametopricedict = collections.defaultdict(list)
        for key, value in zip(productnamelistfordict, productqtylistfordict): #MEEDS TO CHANGE
            productnametoqtydict[key].append(value)
        for key, value in zip(productnamelistfordict, prodictpricelistfordict):   #MEEDS TO CHANGE
            productnametopricedict[key].append(value)
        for y in POtoProductNameDict.get(k):
            SupCodeBox.insert(END,POSupListdict.get(str(k)))   
            POCodeBox.insert(END,k) 
            SKUNameBox.insert(END,y) 
            QtyBox.insert(END,str(productnametoqtydict.get(y))) 
            TotalPriceBox.insert(END,str(productnametopricedict.get(y))) 
        TotalPriceBox.insert(END,sum(POCostListdict.get(k)))     ######  INSERTS FINAL PRICE SUM (NEED TO ADD  SPACE TO EVEN OUT)
        SupCodeBox.insert(END," ")   
        POCodeBox.insert(END," ") 
        SKUNameBox.insert(END," ") 
        QtyBox.insert(END," ") 
    root.update()
    v1.img_object_li.clear()
def select4(event):
    global v1
    v1.img_object_li.clear()
    g =Passlist.get(ANCHOR)
    v2 = v1.pdf_view(frameLeftSide,pdf_location = r"C:\\Users\\nickb\\Desktop\\Python\\"+g, bar=False, width = 122, height = 60) # 100 by 60
    v2.place(x=0, y=0)
    root.update()
    global POSupListdict
    global POCostListdict
    global POtoProductNameDict
    global POtoQtyDict
    global SupCodeBox
    global POCodeBox
    global SKUNameBox
    global QtyBox
    global TotalPriceBox
    SupCodeBox.delete(0,END)
    POCodeBox.delete(0,END)
    SKUNameBox.delete(0,END)
    QtyBox.delete(0,END)
    TotalPriceBox.delete(0,END)
    for k in InvoiceNametoInvoicePONumbDict.get(g):
        productnamelistfordict = []
        productqtylistfordict = []
        prodictpricelistfordict = []
        for b in POtoProductNameDict.get(k):
            productnamelistfordict.append(b)
        for w in POtoQtyDict.get(k):
            productqtylistfordict.append(w)
        for q in POCostListdict.get(k):
            prodictpricelistfordict.append(q)
        productnametoqtydict = collections.defaultdict(list)
        productnametopricedict = collections.defaultdict(list)
        for key, value in zip(productnamelistfordict, productqtylistfordict): #MEEDS TO CHANGE
            productnametoqtydict[key].append(value)
        for key, value in zip(productnamelistfordict, prodictpricelistfordict):   #MEEDS TO CHANGE
            productnametopricedict[key].append(value)
        for y in POtoProductNameDict.get(k):
            SupCodeBox.insert(END,POSupListdict.get(str(k)))   
            POCodeBox.insert(END,k) 
            SKUNameBox.insert(END,y) 
            QtyBox.insert(END,str(productnametoqtydict.get(y))) 
            TotalPriceBox.insert(END,str(productnametopricedict.get(y))) 
        TotalPriceBox.insert(END,sum(POCostListdict.get(k)))     ######  INSERTS FINAL PRICE SUM (NEED TO ADD  SPACE TO EVEN OUT)
        SupCodeBox.insert(END," ")   
        POCodeBox.insert(END," ") 
        SKUNameBox.insert(END," ") 
        QtyBox.insert(END," ") 
    root.update()
    v1.img_object_li.clear()
def select5(event):
    global v1
    v1.img_object_li.clear()
    g =Passlist.get(ANCHOR)
    v2 = v1.pdf_view(frameLeftSide,pdf_location = r"C:\\Users\\nickb\\Desktop\\Python\\"+g, bar=False, width = 122, height = 60) # 100 by 60
    v2.place(x=0, y=0)
    root.update()
    global POSupListdict
    global POCostListdict
    global POtoProductNameDict
    global POtoQtyDict
    global SupCodeBox
    global POCodeBox
    global SKUNameBox
    global QtyBox
    global TotalPriceBox
    SupCodeBox.delete(0,END)
    POCodeBox.delete(0,END)
    SKUNameBox.delete(0,END)
    QtyBox.delete(0,END)
    TotalPriceBox.delete(0,END)
    for k in InvoiceNametoInvoicePONumbDict.get(g):
        productnamelistfordict = []
        productqtylistfordict = []
        prodictpricelistfordict = []
        for b in POtoProductNameDict.get(k):
            productnamelistfordict.append(b)
        for w in POtoQtyDict.get(k):
            productqtylistfordict.append(w)
        for q in POCostListdict.get(k):
            prodictpricelistfordict.append(q)
        productnametoqtydict = collections.defaultdict(list)
        productnametopricedict = collections.defaultdict(list)
        for key, value in zip(productnamelistfordict, productqtylistfordict): #MEEDS TO CHANGE
            productnametoqtydict[key].append(value)
        for key, value in zip(productnamelistfordict, prodictpricelistfordict):   #MEEDS TO CHANGE
            productnametopricedict[key].append(value)
        for y in POtoProductNameDict.get(k):
            SupCodeBox.insert(END,POSupListdict.get(str(k)))   
            POCodeBox.insert(END,k) 
            SKUNameBox.insert(END,y) 
            QtyBox.insert(END,str(productnametoqtydict.get(y))) 
            TotalPriceBox.insert(END,str(productnametopricedict.get(y))) 
        TotalPriceBox.insert(END,sum(POCostListdict.get(k)))     ######  INSERTS FINAL PRICE SUM (NEED TO ADD  SPACE TO EVEN OUT)
        SupCodeBox.insert(END," ")   
        POCodeBox.insert(END," ") 
        SKUNameBox.insert(END," ") 
        QtyBox.insert(END," ") 
    root.update()
    v1.img_object_li.clear()
def SortFiles():
    try:
        ExcelFile = openpyxl.load_workbook(filename)
    except:
        FileOpened.configure(text = "ERROR: no file selected")
        return
    #########################
    global passedcounter
    global InvoiceBeingProcessed
    global issuecounter
    global Notonlogcounter
    global ToCheckCounter
    global tocheckinvoies
    global ListofPOstobePassed
    global totalInvoicesInFolder
    ##########################
    global POSupListdict
    global POCostListdict
    global POtoProductNameDict
    global POtoQtyDict
    ##########################
    passedcounter = 0
    issuecounter = 0
    Notonlogcounter = 0
    InvoiceBeingProcessed = 0
    ToCheckCounter = 0
    TotalPassed.configure(text="Amount of passed: "+str(passedcounter))
    TotalIssuesFound.configure(text="Amount of issues: "+str(issuecounter))
    TotalNotOnLog.configure(text="Amount not on log: "+str(Notonlogcounter))
    TotalNoPoFound.configure(text="Amount to check: "+str(ToCheckCounter))
    TotalFilesProcessedProcessing.configure(text = "")
    root.update()
    ExcelSheet = ExcelFile.active
    POListInExcel = []
    SupCodeInExcel = []
    PriceCostInExcel = []
    ProductNamesinExcel = []
    ProductQtyinExcel = []
    for cell in ExcelSheet['A']:
        SupCodeInExcel.append(cell.value)
    for cell in ExcelSheet['C']:
        POListInExcel.append(cell.value)
    for cell in ExcelSheet['E']:
        PriceCostInExcel.append(cell.value)
    for cell in ExcelSheet['B']:                      ###NEED TO CHANGE
        ProductNamesinExcel.append(cell.value)
    for cell in ExcelSheet['D']:                      ###NEED TO CHANGE
        ProductQtyinExcel.append(cell.value)
    POSupListdict = dict(zip(POListInExcel, SupCodeInExcel))
    POCostListdict = collections.defaultdict(list)
    POtoProductNameDict = collections.defaultdict(list)
    POtoQtyDict = collections.defaultdict(list)
    for key, value in zip(POListInExcel, PriceCostInExcel):
        POCostListdict[key].append(value)
    for key, value in zip(POListInExcel, ProductNamesinExcel): #MEEDS TO CHANGE
        POtoProductNameDict[key].append(value)
    for key, value in zip(POListInExcel, ProductQtyinExcel):   #MEEDS TO CHANGE
        POtoQtyDict[key].append(value)
    for Invoices in InvoicesInFolderList:
        InvoiceBeingProcessed += 1
        TotalFilesProcessedProcessing.configure(text = "Processing file %d/%d" % (InvoiceBeingProcessed,totalInvoicesInFolder))
        root.update()
        InvoicesNoForName = []
        NotIssue = 0
        NotOnLog = 0
        PoteNumberList = []
        pdfFile = open(Invoices,'rb')
        try:
            pdfReader = PyPDF3.PdfFileReader(pdfFile)
            NumberOfPages = pdfReader.numPages
            allpages = []
            for pages in range(NumberOfPages):
                page = pdfReader.getPage(pages)
                pageContent = page.extractText()
                for content in pageContent.split():
                    allpages.append(content)         #need to make an all pages for tesseract so Excel price can be searched back into Tesseract
            for PoteNumbers in allpages:
                g = float()
                try:
                    g = int(PoteNumbers)
                except:
                    pass
                if PoteNumbers.startswith('4') and len(PoteNumbers) == 6 and isinstance(g, int) or PoteNumbers.startswith('3') and len(PoteNumbers) == 6 and isinstance(g, int):
                    PoteNumberList.append(PoteNumbers)
        except:
            pass
        if PoteNumberList == []:  #### rather than close, use the tesseract here to build a Pote number list of use (each pdf file will go to tesseract if PDF fails
            allpages = []
            #page = convert_from_path(Invoices, 300,poppler_path = r'C:\\Users\\nbolade\\Desktop\\Main Folder\\Core Files\\poppler-0.68.0_x86\\poppler-0.68.0\\bin')
            page = convert_from_path(Invoices, 300,poppler_path = r'C:\\Users\\nickb\\Desktop\\Python\\poppler-0.68.0_x86\\poppler-0.68.0\\bin')
            page[0].save(Invoices[:-4]+'.jpg', 'JPEG')
            x = Image.open(Invoices[:-4]+'.jpg')
            pageContent = pytesseract.image_to_string(x)
            for content in pageContent.split():
                allpages.append(content)         #need to make an all pages for tesseract so Excel price can be searched back into Tesseract
            x.close()
            os.remove(Invoices[:-4]+'.jpg')
            for PoteNumbers in allpages:
                g = float()
                try:
                    g = int(PoteNumbers)
                except:
                    pass
                if PoteNumbers.startswith('4') and len(PoteNumbers) == 6 and isinstance(g, int) or PoteNumbers.startswith('3') and len(PoteNumbers) == 6 and isinstance(g, int):
                    PoteNumberList.append(PoteNumbers)
            if PoteNumberList == []:
                pdfFile.close()
                ToCheckCounter += 1
                TotalNoPoFound.configure(text="Amount to check: "+str(ToCheckCounter))
                tocheckinvoies.append(Invoices)
                root.update()
                continue
        PoteNumbersThatHaveMatched = []
        for numbers in PoteNumberList:
            if numbers in POListInExcel:
                if numbers in PoteNumbersThatHaveMatched:
                    continue
                PoteNumbersThatHaveMatched.append(numbers)
                TotalPoCost = sum(POCostListdict.get(numbers))
                InvoicesNoForName.append(numbers)
                NotIssue = 1
                if str(f'{TotalPoCost:.2f}') in allpages or str(f'{TotalPoCost:,.2f}') in allpages or str('£'+f'{TotalPoCost:,.2f}') in allpages or str('£'+f'{TotalPoCost:.2f}') in allpages:  #make sure to switch all pages back to Tesseract
                    NotIssue = 2
            else: 
                NotOnLog = 1
        if PoteNumbersThatHaveMatched != []:
            InvoiceNametoInvoicePONumbDict[Invoices] = PoteNumbersThatHaveMatched
        if NotIssue == 2:
            pdfFile.close()
            passedcounter += 1
            TotalPassed.configure(text="Amount of passed: "+str(passedcounter))
            ListofPOstobePassed.append(Invoices)
            Passlist.insert(END,Invoices)
            root.update()
        elif NotIssue == 1:
            pdfFile.close()
            issuecounter += 1
            TotalIssuesFound.configure(text="Amount of issues: "+str(issuecounter))
            IssueList.insert(END,Invoices)
            root.update()
        elif NotOnLog == 1:
            pdfFile.close()
            Notonlogcounter += 1
            TotalNotOnLog.configure(text="Amount not on log: "+str(Notonlogcounter))
            NoPOFoundList.insert(END,Invoices)
            root.update()
            continue
    TotalFilesProcessedProcessing.configure(text = "All files proccessed %d/%d" % (totalInvoicesInFolder,totalInvoicesInFolder))
    TotalPassed.configure(text="Amount of passed: "+str(passedcounter))
    TotalIssuesFound.configure(text="Amount of issues: "+str(issuecounter))
    TotalNotOnLog.configure(text="Amount not on log: "+str(Notonlogcounter))
    TotalNoPoFound.configure(text="Amount to check: "+str(ToCheckCounter))
    root.update()
###############################################################################
for i in range(99):   ### 20 rows/columns each
    root.columnconfigure(i,weight=1)
    root.rowconfigure(i,weight=1)
###############################################################################
frameRightSide = Frame(root)
for i in range(79): 
    frameRightSide.columnconfigure(i,weight=1)
    frameRightSide.rowconfigure(i,weight=1)
frameRightSide.grid(row=1, column=90, columnspan=8, rowspan=97,sticky="NSEW")
frameRightSide1 = Frame(root, bg="grey")
for i in range(79): 
    frameRightSide1.columnconfigure(i,weight=1)
    frameRightSide1.rowconfigure(i,weight=1)
frameRightSide1.grid(row=1, column=80, columnspan=8, rowspan=97,sticky="NSEW")
frameLeftSide = Frame(root,bg="darkgrey")
frameLeftSide.grid(row=1, column=1,rowspan=97, columnspan=35,sticky="NSEW")  #### edit column span to enlarge pdfviwer
frameCenterBottom = Frame(root,bg="darkgrey")
for i in range(19):   ### 20 rows/columns each
    frameCenterBottom.columnconfigure(i,weight=1)
    frameCenterBottom.rowconfigure(i,weight=1)
frameCenterBottom.grid(row=50, column=37,rowspan=48, columnspan=35,sticky="NSEW")
framCenterTop1 = Frame(root,bg="darkgrey",relief=SUNKEN,bd=2)
for i in range(19):   ### 20 rows/columns each
    framCenterTop1.columnconfigure(i,weight=1)
    framCenterTop1.rowconfigure(i,weight=1)
framCenterTop1.grid(row=5, column=37,rowspan=26, columnspan=35,sticky="NSEW")
framCenterTop2 = Frame(root,bg="darkgrey",relief=SUNKEN,bd=2)
for i in range(19):   ### 20 rows/columns each
    framCenterTop2.columnconfigure(i,weight=1)
    framCenterTop2.rowconfigure(i,weight=1)
framCenterTop2.grid(row=34, column=37,rowspan=14, columnspan=35,sticky="NSEW")
###############################################################################
v1 = pdf.ShowPdf()
v2 = v1.pdf_view(frameLeftSide,pdf_location = r"C:\\Users\\nickb\\Desktop\\Python\\Test2.pdf", bar=False, width = 122, height = 60) #idth = 150, height = 80
v2.place(x=0, y=0)
#################################
########### GRIDS ################        ########  49 is the halfway Mark for both up and down, with 99 being the last grid
#################################
TotalPassed = Label(framCenterTop2, bg="darkgrey",text="Total Passed:",width=20, height=1)
TotalPassed.grid(column=1, row=6, columnspan=1, rowspan=1,sticky=E)
TotalIssuesFound = Label(framCenterTop2, bg="darkgrey",text="Total Issues:",width=20, height=1)
TotalIssuesFound.grid(column=1, row=7, columnspan=1, rowspan=1,sticky=E)
TotalNoPoFound = Label(framCenterTop2, bg="darkgrey",text="Total No PO Found:",width=20, height=1)
TotalNoPoFound.grid(column=1, row=8, columnspan=1, rowspan=1,sticky=E)
TotalNotOnLog = Label(framCenterTop2, bg="darkgrey",text="Total Not On Log:",width=20, height=1)
TotalNotOnLog.grid(column=1, row=9, columnspan=1, rowspan=1,sticky=E)
#############################################################################################################################################
FileOpened = Label(framCenterTop1, bg="darkgrey",text="File Opened:                       ",width=1, height=1)
FileOpened.grid(column=0, row=6, columnspan=7, rowspan=1, sticky=NSEW)
TotalFilesttoProcess = Label(framCenterTop1, bg="darkgrey",text="Total Files To Process:        ",width=1, height=1).grid(column=0, row=7, columnspan=7, rowspan=1, sticky=NSEW)
TotalFilesProcessedProcessing = Label(framCenterTop1, bg="darkgrey",text="Total Files Processed:         ",width=1, height=1)
TotalFilesProcessedProcessing.grid(column=0, row=8, columnspan=7, rowspan=1, sticky=NSEW)
#############################################################################################################################################
POCodeInputLabel = Label(framCenterTop1,width = 1, text="Search PO Code", bg = 'darkgrey').grid(column=15, row=6, columnspan=2, rowspan=1, sticky=NSEW)
POCodeSearchLabel = Button(framCenterTop1,width = 1, text="Search", bg = 'grey').grid(column=15, row=9, columnspan=2, rowspan=1, sticky=NSEW)
POCodeAssignLabel = Button(framCenterTop1,width = 1, text="Assign", bg = 'grey').grid(column=15, row=11, columnspan=2, rowspan=1, sticky=NSEW)
POCodeInputBox = Entry(framCenterTop1,width = 1, bg = 'lightgrey',bd=3).grid(column=15, row=7, columnspan=2, rowspan=1, sticky=NSEW)
Rb1 = Radiobutton(framCenterTop1,text = "Add to pass",value="3",width = 1,bg = 'darkgrey', height = 1,command=radioselect1).grid(column=15, row=1, columnspan=2, rowspan=1, sticky=NSEW) #variable=selectpass.x
Rb2 = Radiobutton(framCenterTop1,text = "remove from pass",value="4",width = 1,bg = 'darkgrey', height = 1,command=radioselect2).grid(column=15, row=3, columnspan=3, rowspan=1, sticky=NSEW) #variable=selectpass.x
#############################################################################################################################################
button_browser = Button(framCenterTop1,text = "Browse Files",bg = 'grey',width = 1, height = 1, command=browser).grid(column=1, row=1, columnspan=2, rowspan=1, sticky=NSEW)
button_sort = Button(framCenterTop1,text = "Sort",bg = 'grey',width = 10,height = 1, command=SortFiles).grid(column=1, row=3, columnspan=2, rowspan=1, sticky=NSEW)
button_exit = Button(framCenterTop1,text = "Exit",bg = 'grey',width = 10,height = 1).grid(column=4, row=3, columnspan=2, rowspan=1, sticky=NSEW)
button_execute = Button(framCenterTop1,text = "Execute",bg = 'grey',width = 10,height = 1).grid(column=4, row=1, columnspan=2, rowspan=1, sticky=NSEW)
#############################################################################################################################################
PassedListLabel = Label(frameRightSide, bg="grey",text="TO SORT LIST",width=1, height=1).grid(column=0, row=0, columnspan=79, rowspan=1, sticky=NSEW)
Passlist = Listbox(frameRightSide, bg="darkgrey", width=1, height=1, selectmode='single')
Passlist.grid(column=0, row=1, columnspan=79, rowspan=26, sticky=NSEW)
Passlist.bind('<<ListboxSelect>>', select)
IssuesListLabel = Label(frameRightSide, bg="grey",text="ISSUE LIST",width=1, height=1).grid(column=0, row=27, columnspan=79, rowspan=1, sticky=NSEW)
IssueList = Listbox(frameRightSide, bg="darkgrey",width=1, height=1, selectmode='single')
IssueList.grid(column=0, row=28, columnspan=79, rowspan=27, sticky=NSEW)
IssueList.bind('<<ListboxSelect>>', select2)
NoPOFoundLabel = Label(frameRightSide, bg="grey",text="NO PO LIST",width=1, height=1).grid(column=0, row=55, columnspan=79, rowspan=1, sticky=NSEW)
NoPOFoundList = Listbox(frameRightSide, bg="darkgrey",width=1, height=1, selectmode='single')
NoPOFoundList.grid(column=0, row=56, columnspan=79, rowspan=26, sticky=NSEW)
NoPOFoundList.bind('<<ListboxSelect>>', select3)
################################################
UnusedLabel= Label(frameRightSide1, bg="grey",text="UNUSED",width=1, height=1).grid(column=0, row=0, columnspan=79, rowspan=1, sticky=NSEW)
UnusedListbox = Listbox(frameRightSide1, bg="darkgrey", width=1, height=1, selectmode='single')
UnusedListbox.grid(column=0, row=1, columnspan=79, rowspan=26, sticky=NSEW)
UnusedListbox.bind('<<ListboxSelect>>', select)
NotonLogLabel = Label(frameRightSide1, bg="grey",text="NOT ON LOG",width=1, height=1).grid(column=0, row=27, columnspan=79, rowspan=1, sticky=NSEW)
NotOnLogListBox = Listbox(frameRightSide1, bg="darkgrey",width=1, height=1, selectmode='single')
NotOnLogListBox.grid(column=0, row=28, columnspan=79, rowspan=27, sticky=NSEW)
#NotOnLogListBox.bind('<<ListboxSelect>>', select2)
POPassedListLabel = Label(frameRightSide1, bg="grey",text="PASSED LIST",width=1, height=1).grid(column=0, row=55, columnspan=79, rowspan=1, sticky=NSEW)
POPassedListBox = Listbox(frameRightSide1, bg="darkgrey",width=1, height=1, selectmode='single')
POPassedListBox.grid(column=0, row=56, columnspan=79, rowspan=26, sticky=NSEW)
#POPassedListBox.bind('<<ListboxSelect>>', select3)
############################################################################################################################################
SupCodeLabel = Label(frameCenterBottom, bg="lightgrey",text="SUP",width=1, height=1).grid(column=0, row=0, columnspan=3, rowspan=1, sticky=NSEW)
SupCodeBox = Listbox(frameCenterBottom, bg="darkgrey",width=1, height=1)
SupCodeBox.grid(column=0, row=1, columnspan=3, rowspan=18, sticky=NSEW)
PoCodeLabel = Label(frameCenterBottom, bg="lightgrey",text="PO",width=1, height=1).grid(column=3, row=0, columnspan=2, rowspan=1, sticky=NSEW)
POCodeBox = Listbox(frameCenterBottom, bg="darkgrey",width=1, height=1)
POCodeBox.grid(column=3, row=1, columnspan=2, rowspan=18, sticky=NSEW)
SUKNameLabel = Label(frameCenterBottom, bg="lightgrey",text="SKU PRODUCT NAME",width=1, height=1).grid(column=5, row=0, columnspan=8, rowspan=1, sticky=NSEW)
SKUNameBox = Listbox(frameCenterBottom, bg="darkgrey",width=1, height=1)
SKUNameBox.grid(column=5, row=1, columnspan=8, rowspan=18, sticky=NSEW)
QtyLabel = Label(frameCenterBottom, bg="lightgrey",text="QTY",width=1, height=1).grid(column=13, row=0, columnspan=3, rowspan=1, sticky=NSEW)
QtyBox = Listbox(frameCenterBottom, bg="darkgrey",width=1, height=1)
QtyBox.grid(column=13, row=1, columnspan=3, rowspan=18, sticky=NSEW)
PriceLabel = Label(frameCenterBottom, bg="lightgrey",text="PRICE",width=1, height=1).grid(column=16, row=0, columnspan=3, rowspan=1, sticky=NSEW)
TotalPriceBox = Listbox(frameCenterBottom, bg="darkgrey",width=1, height=1)
TotalPriceBox.grid(column=16, row=1, columnspan=3, rowspan=18, sticky=NSEW)
############################################################################################################################################


root.mainloop()