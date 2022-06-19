###### T89 EMAIL SENDER
## TO DO 
## MAKE SO THAT ALL THE FOLDERS IN THE TWO SEND FILE GET SENT(WHETHER PROCESSED OR NOT)
## ADD SHUTIL MOVE SO FILES THAT ARE MOVED TO THE SENDING FOLDER ARE REMOVED
## ADD ANOTHER SHUTIL MOVE SO THAT FILES SENT ARE ALSO MOVED TO THAT FOLDER
## DO MORE TESTING!!!
## COMPILE AND MAKE IT WORKS ON INDEPENDENT LAPTOP (IT FAILED TO SORT ON KARENS, FIGURE OUT WHY...)
## COMPILE IT ON WORKDESKTOP AND TRY IN HOMEPC
### need to set automatic creation of ready to send folder and sent folder. 
from tkinter import *
from tkinter import filedialog
from pathlib import Path
import win32com.client as win32
import pytesseract, os, PyPDF3, openpyxl, sys, shutil
from openpyxl.utils.cell import get_column_letter
from pdf2image import convert_from_path
from PIL import ImageTk, Image
import time
MainWindow=Tk()
MainWindow.title("Sending Emails")
MainWindow.geometry("600x500+600+200")
MainWindow.configure(bg="RoyalBlue4")
def T89Selected():
    PopplerPath = os.getcwd()+'\\poppler-0.68.0_x86\\poppler-0.68.0\\bin'
    Tesspath = os.getcwd()+'\\Tesseract-OCR\\tesseract.exe'
    poppler_path = PopplerPath
    pytesseract.pytesseract.tesseract_cmd = Tesspath
    T89SenderWindow=Toplevel(MainWindow)
    T89SenderWindow.title("Sending Emails")
    T89SenderWindow.geometry("900x800+500+100")
    T89SenderWindow.configure(bg="RoyalBlue4")
    T89SenderWindow.attributes('-topmost',1)
    global Safety
    global SortSafety1
    global SortSafety2
    global SortSafety3
    global T89sInFolderList
    global T89stoSend
    global T89SentCOunter
    global FilesRemaned
    T89sInFolderList = []
    T89stoSend = []
    totalT89sInFolder = 0
    T89SentCOunter = 0
    FilesRemaned = 0
    Safety = 'On'
    SortSafety1 = 'On'
    SortSafety2 = 'On'
    SortSafety3 = 'Off'
    def browser():
        labelforERRORs.configure(text='')
        global filename
        global ExcelFile
        global ExcelSheet
        global EmailAddressList
        global PayrollNumberList
        global NamesList
        global PaynumberToEmailDict
        global EmailtoNameDict
        global OKJKInNumberList
        global POtoOKJKIndict
        global NamesList2List
        global Name1toName2Dict
        global SortSafety1
        global NamesList2ListtoPayNumber
        NamesList2ListtoPayNumber = []
        EmailAddressList = []
        PayrollNumberList = []
        NamesList = []
        OKJKInNumberList = []
        NamesList2List = []
        try:
            filename = filedialog.askopenfilename(parent=T89SenderWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select file",filetypes = (("Excel","*xlsx"),))
            ExcelFile = openpyxl.load_workbook(filename)
            ExcelSheet = ExcelFile.active
            filename0 = Path(filename).stem 
            labelfileopened.configure(text="File Opened: "+filename0)
        except:
            labelforERRORs.configure(text='Please select a valid excel file')
            SortSafety1 = 'On'
            return
        for cell in ExcelSheet['I']:
            EmailAddressList.append(cell.value)
        for cell in ExcelSheet['C']:
            PayrollNumberList.append(cell.value)
        for cell in ExcelSheet['D']:
            OKJKInNumberList.append(cell.value)    
        for cell in ExcelSheet['B']:
            NamesList.append(cell.value)
        for cell in ExcelSheet['E']:
            NamesList2List.append(cell.value)
        PaynumberToEmailDict = dict(zip(PayrollNumberList, EmailAddressList))
        EmailtoNameDict = dict(zip(EmailAddressList, NamesList))
        Name1toName2Dict = dict(zip(NamesList, NamesList2List))
        POtoOKJKIndict = dict(zip(PayrollNumberList, OKJKInNumberList))
        NamesList2ListtoPayNumber = dict(zip(NamesList2List, PayrollNumberList))
        SortSafety1 = 'Off'
        T89SenderWindow.update()
    def browser2():
        global SortSafety2
        global SortSafety3
        labelforERRORs.configure(text="")
        global T89sInFolderList
        global totalT89sInFolder
        global T89Folder
        T89sInFolderList = []
        T89Folder = filedialog.askdirectory(parent=T89SenderWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select T89 Folder")
        try:
            T89sInFolder = os.listdir(T89Folder)
        except:
            labelforERRORs.configure(text="ERROR: No folder selected")
            return
        SortSafety3 = 'Off'
        if 'T89s Ready To Send' in T89Folder[-18:] or 'T89s Sent' in T89Folder[-9:]:
            labelforERRORs.configure(text="ERROR: Please do not select the 'T89s To Send' folder or subfolders")
            T89SenderWindow.update()
            return
        for files in T89sInFolder:
            if files.endswith('.pdf') or files.endswith('.PDF'):
                T89sInFolderList.append(files)
        totalT89sInFolder = len(T89sInFolderList)
        if totalT89sInFolder == 0:
            labelforERRORs.configure(text="ERROR: No PDF files found in folder, please select correct folder")
            T89SenderWindow.update()
            return
        window1 = Toplevel(T89SenderWindow)
        window1.title('All PDF files in folder, please select another folder if wrong')
        window1.geometry("900x700+500+200") 
        window1.configure(bg="RoyalBlue4")
        window1.attributes('-topmost',1)
        ListofPDFfilesinFolder = Listbox(window1, bg="grey",width=80, height=31, selectmode='single', font=('Times', 14))
        ListofPDFfilesinFolder.pack()
        ListofPDFfilesinFolder.insert(END,'All PDF files in selected folder, please choose another folder if wrong')
        ListofPDFfilesinFolder.insert(END,'')
        ListofPDFfilesinFolder.insert(END,'Total files in folder: '+str(totalT89sInFolder))
        ListofPDFfilesinFolder.insert(END,'')
        for x in T89sInFolderList:
            ListofPDFfilesinFolder.insert(END,x)
        labetotalT89tosend.configure(text="T89s to be processed: "+str(totalT89sInFolder))
        try:
            if len(os.listdir(T89Folder+'\\T89s Ready To Send')) > 1:
                labelforERRORs.configure(text="Warning, files found in 'T89s ToSend' folder, please delete these files before sorting")
                SortSafety3 = 'On'
                T89SenderWindow.update()
                return
            else:
                SortSafety2 = 'Off'
                T89SenderWindow.update()
        except:
            SortSafety2 = 'Off'
            T89SenderWindow.update()
            pass
    def Sort():
        if  SortSafety3 == 'On':
            labelforERRORs.configure(text="Warning, files found in 'T89s ToSend' folder, please delete these files before sorting")
            T89SenderWindow.update()
            return
        if SortSafety1 == 'On':
            labelforERRORs.configure(text="Please select a leavers report before sorting")
            T89SenderWindow.update()
            return
        if SortSafety2 == 'On':
            labelforERRORs.configure(text="Please select the T89 folder before sorting")
            T89SenderWindow.update()
            return
        ListofT89stoEmailAddress = []
        Path(T89Folder+'\\T89s Ready To Send\\T89s Sent').mkdir(parents=True, exist_ok=True)
        Path(T89Folder+'\\T89s Ready To Send').mkdir(parents=True, exist_ok=True)
        labelforERRORs.configure(text="")
        PayrollNumber = ""
        EmailAddress = "" 
        Name = ""
        labetotalT89tosend.configure(text="")
        global T89sInFolderList
        global FilesRemaned
        global T89stoSend
        global Safety
        global totalT89sInFolder
        Safety = 'On'
        T89stoSend = []
        FilesRemaned = 0
        Error = False
        for T89s in T89sInFolderList:
            PayrollNumber = ""
            EmailAddress = ""
            Name = ""
            OKJKNumber = ""
            newFileName = ''
            nameForRename = ''
            page = convert_from_path(T89Folder+"\\"+T89s, 600,poppler_path = PopplerPath) 
            page[0].save(T89s[:-4]+'.jpg', 'JPEG')
            x = Image.open(T89s[:-4]+'.jpg')
            pageContent = pytesseract.image_to_string(x)
            for content in pageContent.split():
                if len(content) == 6 and content.startswith("0"): ## and is int
                    PayrollNumber = content
            if PayrollNumber == "":
                labelforERRORs.configure(text="ERROR: Failed to find payroll number in file: "+T89s+". Please run a new report and sort again")
                Error = True
                x.close()
                os.remove(T89s[:-4]+'.jpg')
                break
            EmailAddress = PaynumberToEmailDict.get(PayrollNumber)
            Name = EmailtoNameDict.get(EmailAddress)
            nameForRename = Name1toName2Dict.get(Name)
            OKJKNumber =  POtoOKJKIndict.get(PayrollNumber)
            if Name in pageContent.split():
                pass
            else:
                labelforERRORs.configure(text="ERROR: Name does not match report: "+T89s+". Please run a new report and sort again")
                x.close()
                os.remove(T89s[:-4]+'.jpg')
                Error = True
                break
            x.close()
            os.remove(T89s[:-4]+'.jpg')
            PDFFile = PyPDF3.PdfFileReader(T89Folder+'/'+T89s)
            NumberOfPages = PDFFile.numPages
            Output_PDFFile = PyPDF3.PdfFileWriter()
            for i in range(NumberOfPages):
                Output_PDFFile.addPage(PDFFile.getPage(i))
            Output_PDFFile.encrypt(OKJKNumber)
            Output_PDFFile.write(open(T89Folder+'\\T89s Ready To Send\\'+PayrollNumber+" "+nameForRename+".pdf", 'wb'))
            FilesRemaned += 1
            newFileName = ''
            OKJKNumber = ""
            PayrollNumber = ""
            EmailAddress = ""
            Name = ""
            nameForRename = ''
            labetotalT89tosend.configure(text='File being Processed: '+str(FilesRemaned)+'/'+str(totalT89sInFolder))
            T89SenderWindow.update()
        T89stoSend = os.listdir(T89Folder+'\\T89s Ready To Send')
        for files4 in T89stoSend:
            if files4.endswith('.pdf') or files4.endswith('.PDF'):
                newFileName = ''
                EmailAddress = ""
                PayrollNumber = ''
                NameOnthePDFFile = ''
                files1 = ''
                files1 = Path(files4).stem 
                PayrollNumber = files1[:6]
                NameOnthePDFFile = files1[7:]
                newFileName = PayrollNumber+' '+NameOnthePDFFile
                EmailAddress = PaynumberToEmailDict.get(PayrollNumber)
                OKJKNumber =  POtoOKJKIndict.get(PayrollNumber)
                ListofT89stoEmailAddress.append(newFileName+'    ---->    '+EmailAddress)
                files1 = ''
                PayrollNumber = ''
                NameOnthePDFFile = ''
                newFileName = ''
                EmailAddress = ""  
        if Error == False:
            labetotalT89tosend.configure(text="Total ready to send is "+str(len(T89stoSend)-1))
            window = Toplevel(T89SenderWindow)
            window.title('List of T89s and Address for them to be sent to')
            window.geometry("900x700+500+200") 
            window.configure(bg="RoyalBlue4")
            window.attributes('-topmost',1)
            ListofT89toEmail = Listbox(window, bg="grey",width=80, height=31, selectmode='single', font=('Times', 14))
            ListofT89toEmail.pack()
            ListofT89toEmail.insert(END,'Total T89s to be sent '+str(len(T89stoSend)-1))
            ListofT89toEmail.insert(END,'')
            for x in ListofT89stoEmailAddress:
                ListofT89toEmail.insert(END,x)
            Safety = 'Off'
            T89SenderWindow.update()
        else:
            pass
        ListofT89stoEmailAddress = []
    def Send():
        global Safety
        global T89stoSend
        if Safety == 'Off':
            T89stoSend = os.listdir(T89Folder+'\\T89s Ready To Send')
            PayrollNumber = ""
            EmailAddress = ""
            NameOnthePDFFile = ""
            PayrollNumberCheck = ""
            Name = ""
            global FilesRemaned
            global T89SentCOunter
            global totalT89sInFolder
            for T89stos in T89stoSend:
                if T89stos.endswith('.pdf') or T89stos.endswith('.PDF'):
                    PayrollNumber = ""
                    EmailAddress = ""
                    NameOnthePDFFile = ""
                    PayrollNumberCheck = ""
                    Name = ""
                    OKJKNumber = ""
                    nameForRename = ""
                    file = Path(T89Folder+'\\T89s Ready To Send\\'+T89stos).stem           #### CHANGE FOLDER NAME HERE 
                    PayrollNumber = file[:6]
                    NameOnthePDFFile = file[7:]
                    EmailAddress = PaynumberToEmailDict.get(PayrollNumber)
                    PayrollNumberCheck = NamesList2ListtoPayNumber.get(NameOnthePDFFile)
                    if PayrollNumber == PayrollNumberCheck:
                        pass
                    else:
                        labelforERRORs.configure(text="Name of T89 does not match name on file when trying to send "+T89stos+". Please run a new report and sort again")
                        Safety = 'On'
                        break
                    outlook = win32.Dispatch('outlook.application')
                    mail = outlook.CreateItem(0)
                    mail.To = EmailAddress
                    mail.Subject = 'T89'
                    mail.Body = 'Please find your T89 enclosed in this email.\n\nThe password to open your T89, is the same password to open your wageslip. Please contract your old store manager if you are unsure on what this is, as we cannot give hints for the password over email.\n\nRegards\n\nTOFS'
                    mail.Attachments.Add(T89Folder+'\\T89s Ready To Send\\'+T89stos)
                    #mail.CC = "payroll@tofs.com"
                    mail.Send()
                    T89SentCOunter += 1
                    labetotalT89tosend.configure(text = "Amount of files sent: %d/%d" % (T89SentCOunter,totalT89sInFolder))
                    OKJKNumber = ""
                    PayrollNumber = ""
                    EmailAddress = ""
                    NameOnthePDFFile = ""
                    PayrollNumberCheck = ""
                    nameForRename = ""
                    Name = ""
                    shutil.move(T89Folder+'\\T89s Ready To Send\\'+T89stos,T89Folder+'\\T89s Ready To Send\\T89s Sent\\'+T89stos)
                    T89SenderWindow.update()
                    time.sleep(1)
            labetotalT89tosend.configure(text = "All emails sent %d/%d\n\nPlease remove all files in 'T89s To Send', ready for next process" % (T89SentCOunter,totalT89sInFolder))
            T89SenderWindow.update()
            T89SentCOunter = 0
        else:
            labelforERRORs.configure(text='Please sort the T89s before attempting to send')
    label_file_explorer = Label(T89SenderWindow,text = "Original Factory Shop Email Sender - By Nick",width = 100, height = 4,fg = "white",bg = 'RoyalBlue4', font=('Times', 13)) #RoyalBlue4
    button_explore = Button(T89SenderWindow,text = "Select Chris21 Leavers Report",bg = 'RoyalBlue4',width = 30, height = 2,command = browser, font=('Times', 15, 'bold'), fg = "yellow2")
    button_explore2 = Button(T89SenderWindow,text = "Select T89 Folder",bg = 'RoyalBlue4',width = 30, height = 2,command = browser2, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exit = Button(T89SenderWindow,text = "Exit",bg = 'RoyalBlue4',width = 30,height = 2,command = sys.exit, font=('Times', 15, 'bold'), fg = "yellow2")
    button_sort = Button(T89SenderWindow,text = "Sort",bg = 'RoyalBlue4',width = 30,height = 2,command = Sort, font=('Times', 15, 'bold'), fg = "yellow2")
    button_send = Button(T89SenderWindow,text = "Send",bg = 'RoyalBlue4',width = 30,height = 2,command = Send, font=('Times', 15, 'bold'), fg = "yellow2")
    labelfileopened = Label(T89SenderWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labetotalT89tosend = Label(T89SenderWindow,text = "",width = 50, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelforERRORs = Label(T89SenderWindow,text = "",width = 75, height = 2,fg = "red",bg = 'RoyalBlue4', font=('Times', 16))
    labetotalT89tosend = Label(T89SenderWindow,text = "",width = 75,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    TitleImage = os.getcwd()+'\\1519797862804.jpg'
    img = Image.open(TitleImage)
    img = img.resize((500, 100), Image.LANCZOS)
    img = ImageTk.PhotoImage(img)
    labelimage = Label(T89SenderWindow, image = img,width = 500, height = 100)
    def allpacks():
        labelimage.pack()
        label_file_explorer.pack()
        button_explore.pack()
        button_explore2.pack()
        button_sort.pack()
        button_exit.pack()
        button_send.pack()
        labelforERRORs.pack()
        labelfileopened.pack()
        labetotalT89tosend.pack()
        labetotalT89tosend.pack()
    allpacks()
    T89SenderWindow.mainloop()
def NESTSelected():
    NESTWindow=Toplevel(MainWindow)
    NESTWindow.title("Sending NEST emails")
    NESTWindow.geometry("900x800+500+100")
    NESTWindow.configure(bg="RoyalBlue4")
    NESTWindow.attributes('-topmost',1)
    nestlettersent = 0
    global NESTSendSafety1
    global NESTSendSafety2
    NESTSendSafety1 = 'On'
    NESTSendSafety2 = 'On'
    def NestExcelFile():
        global NESTExcelFIle
        global ExcelFileNEST
        global ExcelSheetNEST
        global AddressListNEST
        global TotalEmailAdressesNEST
        global NESTSendSafety1
        NESTSendSafety1 = 'On'
        labelERRORsfoRNEST.configure(text='')
        AddressListNEST = []
        NESTExcelFIle = filedialog.askopenfilename(parent=NESTWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select file",filetypes = (("Excel","*xlsx"),("all files","*.*")))
        labelfileopnedNEST.configure(text="File Opened: "+NESTExcelFIle)
        ExcelFileNEST = openpyxl.load_workbook(NESTExcelFIle)
        ExcelSheetNEST = ExcelFileNEST.active
        for cell in ExcelSheetNEST['N']:
            if cell.value != None and '@' in cell.value:
                AddressListNEST.append(cell.value)
        TotalEmailAdressesNEST = len(AddressListNEST)
        if TotalEmailAdressesNEST > 0:
            pass
        else:
            labelERRORsfoRNEST.configure(text='ERROR: No email address found in Col "N"')
            NESTWindow.update()
            NESTSendSafety1 = 'On'
            return
        NESTWindo1 = Toplevel(NESTWindow)
        NESTWindo1.title('All PDF files in folder, please select another folder if wrong')
        NESTWindo1.geometry("900x700+500+200") 
        NESTWindo1.configure(bg="RoyalBlue4")
        NESTWindo1.attributes('-topmost',1)
        ListofEmailAddressListBoxNEST = Listbox(NESTWindo1, bg="grey",width=80, height=31, selectmode='single', font=('Times', 14))
        ListofEmailAddressListBoxNEST.pack()
        ListofEmailAddressListBoxNEST.insert(END,'All Email Addressses in selected excel file, please choose another file if wrong')
        ListofEmailAddressListBoxNEST.insert(END,'')
        ListofEmailAddressListBoxNEST.insert(END,'Total emails to send to: '+str(TotalEmailAdressesNEST))
        ListofEmailAddressListBoxNEST.insert(END,'')
        for xNEST in AddressListNEST:
            ListofEmailAddressListBoxNEST.insert(END,xNEST)
        labelsentcounterNEST.configure(text="Total Email Addresses to send to: "+str(TotalEmailAdressesNEST))
        NESTSendSafety1 = 'Off'
    def SelectNESTDocx():
        global NESTSendSafety2
        global enrollmentletterNEST
        NESTSendSafety2 = 'On'
        enrollmentletterNEST = ''
        labelERRORsfoRNEST.configure(text='')
        try:
            enrollmentletterNEST = filedialog.askopenfilename(parent=NESTWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select file",filetypes = (("DOCx","*docx"),))
            filename0NEST = Path(enrollmentletterNEST).stem 
            labelfileopnedDocxNEST.configure(text="File To Be Sent: "+filename0NEST)
            NESTSendSafety2 = 'Off'
        except:
            labelERRORsfoRNEST.configure(text='Please select a valid DOCx file')
            return
    def SendNEST():
        global NESTSendSafety1
        global NESTSendSafety2
        if NESTSendSafety1 == 'On':
            labelERRORsfoRNEST.configure(text='Please select Excel file before sending')
            NESTWindow.update()
            return
        if NESTSendSafety2 == 'On':
            labelERRORsfoRNEST.configure(text='Please select DOCx before sending')
            NESTWindow.update()
            return
        else:         
            global TotalEmailAdressesNEST
            global nestlettersent
            global enrollmentletterNEST
            nestlettersent = 0
            for NESTEmails in AddressListNEST:
                outlook = win32.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)
                mail.To = NESTEmails
                mail.Subject = 'NEST ENROLLMENT LETTER'
                mail.Body = 'Hi\n\nEnclosed is your NEST enrollment letter for the company pension\n\nKind Regards,\n\nPAYROLL\n\nTOFS'
                #mail.CC = "payroll@tofs.com"
                mail.Attachments.Add(enrollmentletterNEST)
                mail.Send()
                nestlettersent += 1
                labelsentcounterNEST.configure(text = "Amount of files sent: %d/%d" % (nestlettersent,TotalEmailAdressesNEST))
                time.sleep(1)
                NESTWindow.update()
            labelsentcounterNEST.configure(text = "All emails sent %d/%d" % (nestlettersent,TotalEmailAdressesNEST))
            NESTWindow.update()
    label_file_explorerNEST = Label(NESTWindow,text = "Original Factory Shop Email Sender - By Nick",width = 100, height = 4,fg = "white",bg = 'RoyalBlue4', font=('Times', 13))
    button_exploreExcelNEST = Button(NESTWindow,text = "Select Excel File for Sending",bg = 'RoyalBlue4',width = 30, height = 2,command = NestExcelFile, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exploreDOCxNEST = Button(NESTWindow,text = "Select NEST template file to be Sent",bg = 'RoyalBlue4',width = 30, height = 2,command = SelectNESTDocx, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exitNEST = Button(NESTWindow,text = "Exit",bg = 'RoyalBlue4',width = 30,height = 2,command = sys.exit, font=('Times', 15, 'bold'), fg = "yellow2")
    button_sendNEST = Button(NESTWindow,text = "Send",bg = 'RoyalBlue4',width = 30,height = 2,command = SendNEST, font=('Times', 15, 'bold'), fg = "yellow2")
    labelERRORsfoRNEST = Label(NESTWindow,text = "",width = 75, height = 2,fg = "red",bg = 'RoyalBlue4', font=('Times', 16))
    labelfileopnedNEST = Label(NESTWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelsentcounterNEST = Label(NESTWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelfileopnedDocxNEST = Label(NESTWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    TitleImageNEST = os.getcwd()+'\\1519797862804.jpg'
    imgNEST = Image.open(TitleImageNEST)
    imgNEST = imgNEST.resize((500, 100), Image.LANCZOS)
    imgNEST = ImageTk.PhotoImage(imgNEST)
    labelimageNEST = Label(NESTWindow, image = imgNEST,width = 500, height = 100)
    def allpacks():
        labelimageNEST.pack()
        label_file_explorerNEST.pack()
        button_exploreExcelNEST.pack()
        button_exploreDOCxNEST.pack()
        button_exitNEST.pack()
        button_sendNEST.pack()
        labelERRORsfoRNEST.pack()
        labelfileopnedNEST.pack()
        labelfileopnedDocxNEST.pack()
        labelsentcounterNEST.pack()
    allpacks()
    NESTWindow.mainloop()   
TitleImage = os.getcwd()+'\\1519797862804.jpg'  
img = Image.open(TitleImage)
img = img.resize((500, 100), Image.LANCZOS)
img = ImageTk.PhotoImage(img)
labelimage = Label(MainWindow, image = img,width = 500, height = 100)
Labelfill = Label(MainWindow, width = 30, height = 3,bg = 'RoyalBlue4')
Labelfill1 = Label(MainWindow, width = 30, height = 3,bg = 'RoyalBlue4')
T89WindowButton = Button(MainWindow,text = "Send T89s",bg = 'RoyalBlue4',width = 30,height = 2, font=('Times', 15, 'bold'), fg = "yellow2",command = T89Selected) #command = Sort
NESTWindowButton = Button(MainWindow,text = "Send NEST files",bg = 'RoyalBlue4',width = 30,height = 2, font=('Times', 15, 'bold'), fg = "yellow2",command = NESTSelected) #command = T89Selected
labelimage.pack()
Labelfill.pack()
T89WindowButton.pack()
Labelfill1.pack()
NESTWindowButton.pack()
MainWindow.mainloop()


