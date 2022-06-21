###### P45 EMAIL SENDER
## TO DO 
## MAKE SO THAT ALL THE FOLDERS IN THE TWO SEND FILE GET SENT(WHETHER PROCESSED OR NOT)
## ADD SHUTIL MOVE SO FILES THAT ARE MOVED TO THE SENDING FOLDER ARE REMOVED
## ADD ANOTHER SHUTIL MOVE SO THAT FILES SENT ARE ALSO MOVED TO THAT FOLDER
## DO MORE TESTING!!!
## COMPILE AND MAKE IT WORKS ON INDEPENDENT LAPTOP (IT FAILED TO SORT ON KARENS, FIGURE OUT WHY...)
## COMPILE IT ON WORKDESKTOP AND TRY IN HOMEPC
### need to set automatic creation of ready to send folder and sent folder. 
from tkinter import *
from tkinter import filedialog
from pathlib import Path
import win32com.client as win32
import pytesseract, os, PyPDF3, openpyxl, sys, shutil
from openpyxl.utils.cell import get_column_letter
from pdf2image import convert_from_path
from PIL import ImageTk, Image
import time
MainWindow=Tk()
MainWindow.title("Sending Emails")
MainWindow.geometry("600x500+600+200")
MainWindow.configure(bg="RoyalBlue4")
def P45Selected():
    PopplerPath = os.getcwd()+'\\poppler-0.68.0_x86\\poppler-0.68.0\\bin'
    Tesspath = os.getcwd()+'\\Tesseract-OCR\\tesseract.exe'
    poppler_path = PopplerPath
    pytesseract.pytesseract.tesseract_cmd = Tesspath
    P45SenderWindow=Toplevel(MainWindow)
    P45SenderWindow.title("Sending Emails")
    P45SenderWindow.geometry("900x800+500+100")
    P45SenderWindow.configure(bg="RoyalBlue4")
    P45SenderWindow.attributes('-topmost',1)
    global Safety
    global SortSafety1
    global SortSafety2
    global SortSafety3
    global P45sInFolderList
    global P45stoSend
    global P45SentCOunter
    global FilesRemaned
    P45sInFolderList = []
    P45stoSend = []
    P45SentCOunter = 0
    FilesRemaned = 0
    Safety = 'On'
    SortSafety1 = 'On'
    SortSafety2 = 'On'
    SortSafety3 = 'Off'
    def browser():
        labelforERRORs.configure(text='')
        global filename
        global ExcelFile
        global ExcelSheet
        global EmailAddressList
        global PayrollNumberList
        global NamesList
        global PaynumberToEmailDict
        global EmailtoNameDict
        global NatInNumberList
        global POtoNatIndict
        global NamesList2List
        global Name1toName2Dict
        global SortSafety1
        global SortSafety2
        global NamesList2ListtoPayNumber
        button_explore.configure(bg='red3')
        button_explore2.configure(bg='red3')
        button_sort.configure(bg='red3')
        button_send.configure(bg='red3')
        SortSafety1 = 'On'
        SortSafety2 = 'On'
        NamesList2ListtoPayNumber = []
        EmailAddressList = []
        PayrollNumberList = []
        NamesList = []
        NatInNumberList = []
        NamesList2List = []
        AddressListP45 = []
        try:
            filename = filedialog.askopenfilename(parent=P45SenderWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select file",filetypes = (("Excel","*xlsx"),))
            ExcelFile = openpyxl.load_workbook(filename)
            ExcelSheet = ExcelFile.active
            filename0 = Path(filename).stem 
            labelfileopened.configure(text="File Opened: "+filename0)
        except:
            labelforERRORs.configure(text='Please select a valid excel file')
            SortSafety1 = 'On'
            return
        for cell in ExcelSheet['I']:
            if cell.value != None and '@' in cell.value:
                AddressListP45.append(cell.value)
        TotalEmailAdressesP45 = len(AddressListP45)
        if TotalEmailAdressesP45 > 0:
            AddressListP45 = []
            pass
        else:
            labelforERRORs.configure(text='ERROR: No email address found in Col "I"\nPlease select LEAVERSREPORT2')
            P45SenderWindow.update()
            SortSafety1 = 'On'
            return
        for cell in ExcelSheet['I']:
            EmailAddressList.append(cell.value)
        for cell in ExcelSheet['C']:
            PayrollNumberList.append(cell.value)
        for cell in ExcelSheet['D']:
            NatInNumberList.append(cell.value)    
        for cell in ExcelSheet['B']:
            NamesList.append(cell.value)
        for cell in ExcelSheet['E']:
            NamesList2List.append(cell.value)
        PaynumberToEmailDict = dict(zip(PayrollNumberList, EmailAddressList))
        EmailtoNameDict = dict(zip(EmailAddressList, NamesList))
        Name1toName2Dict = dict(zip(NamesList, NamesList2List))
        POtoNatIndict = dict(zip(PayrollNumberList, NatInNumberList))
        NamesList2ListtoPayNumber = dict(zip(NamesList2List, PayrollNumberList))
        button_explore.configure(bg='green4')
        SortSafety1 = 'Off'
        P45SenderWindow.update()
    def browser2():
        global SortSafety2
        global SortSafety3
        labelforERRORs.configure(text="")
        global P45sInFolderList
        global totalP45sInFolder
        global P45Folder
        button_explore2.configure(bg='red3')
        button_sort.configure(bg='red3')
        button_send.configure(bg='red3')
        SortSafety2 = 'On'
        if SortSafety1 == 'On':
            labelforERRORs.configure(text="Please select a leavers report first")
            return
        P45sInFolderList = []
        P45Folder = filedialog.askdirectory(parent=P45SenderWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select P45 Folder")
        try:
            P45sInFolder = os.listdir(P45Folder)
        except:
            labelforERRORs.configure(text="ERROR: No folder selected")
            return
        SortSafety3 = 'Off'
        if 'P45s Ready To Send' in P45Folder[-18:] or 'P45s Sent' in P45Folder[-9:]:
            labelforERRORs.configure(text="ERROR: Please do not select the 'P45s To Send' folder or subfolders")
            P45SenderWindow.update()
            return
        for files in P45sInFolder:
            if files.endswith('.pdf') or files.endswith('.PDF'):
                P45sInFolderList.append(files)
        totalP45sInFolder = len(P45sInFolderList)
        if totalP45sInFolder == 0:
            labelforERRORs.configure(text="ERROR: No PDF files found in folder, please select correct folder")
            P45SenderWindow.update()
            return
        window1 = Toplevel(P45SenderWindow)
        window1.title('All PDF files in folder, please select another folder if wrong')
        window1.geometry("900x700+500+200") 
        window1.configure(bg="RoyalBlue4")
        window1.attributes('-topmost',1)
        ListofPDFfilesinFolder = Listbox(window1, bg="royalblue1",width=80, height=31, selectmode='single', font=('Times', 14))
        ListofPDFfilesinFolder.pack()
        ListofPDFfilesinFolder.insert(END,'All PDF files in selected folder, please choose another folder if wrong')
        ListofPDFfilesinFolder.insert(END,'')
        ListofPDFfilesinFolder.insert(END,'Total files in folder: '+str(totalP45sInFolder))
        ListofPDFfilesinFolder.insert(END,'')
        for x in P45sInFolderList:
            ListofPDFfilesinFolder.insert(END,x)
        labetotalP45tosend.configure(text="P45s to be processed: "+str(totalP45sInFolder))
        try:
            if len(os.listdir(P45Folder+'\\P45s Ready To Send')) > 1:
                labelforERRORs.configure(text="Warning, files found in 'P45s ToSend' folder, please delete these files before sorting")
                SortSafety3 = 'On'
                P45SenderWindow.update()
                return
            else:
                SortSafety2 = 'Off'
                button_explore2.configure(bg='green4')
                P45SenderWindow.update()
        except:
            SortSafety2 = 'Off'
            button_explore2.configure(bg='green4')
            P45SenderWindow.update()
            pass
    def Sort():
        global P45sInFolderList
        global FilesRemaned
        global P45stoSend
        global Safety
        global totalP45sInFolder
        Safety = 'On'
        button_sort.configure(bg='red3')
        button_send.configure(bg='red3')
        if  SortSafety3 == 'On':
            labelforERRORs.configure(text="Warning, files found in 'P45s ToSend' folder, please delete these files before sorting")
            P45SenderWindow.update()
            return
        if SortSafety1 == 'On':
            labelforERRORs.configure(text="Please select a leavers report before sorting")
            P45SenderWindow.update()
            return
        if SortSafety2 == 'On':
            labelforERRORs.configure(text="Please select the P45 folder before sorting")
            P45SenderWindow.update()
            return
        ListofP45stoEmailAddress = []
        Path(P45Folder+'\\P45s Ready To Send\\P45s Sent').mkdir(parents=True, exist_ok=True)
        Path(P45Folder+'\\P45s Ready To Send').mkdir(parents=True, exist_ok=True)
        labelforERRORs.configure(text="")
        PayrollNumber = ""
        EmailAddress = "" 
        Name = ""
        labetotalP45tosend.configure(text="")
        P45stoSend = []
        FilesRemaned = 0
        Error = False
        for P45s in P45sInFolderList:
            PayrollNumber = ""
            EmailAddress = ""
            Name = ""
            NatNumber = ""
            newFileName = ''
            nameForRename = ''
            page = convert_from_path(P45Folder+"\\"+P45s, 600,poppler_path = PopplerPath) 
            page[0].save(P45s[:-4]+'.jpg', 'JPEG')
            x = Image.open(P45s[:-4]+'.jpg')
            pageContent = pytesseract.image_to_string(x)
            for content in pageContent.split():
                if len(content) == 6 and content.startswith("0"): ## and is int
                    PayrollNumber = content
            if PayrollNumber == "":
                labelforERRORs.configure(text="ERROR: Failed to find payroll number in file: "+P45s+". Please run a new report and sort again")
                Error = True
                x.close()
                os.remove(P45s[:-4]+'.jpg')
                break
            EmailAddress = PaynumberToEmailDict.get(PayrollNumber)
            Name = EmailtoNameDict.get(EmailAddress)
            nameForRename = Name1toName2Dict.get(Name)
            NatNumber =  POtoNatIndict.get(PayrollNumber)
            # if NatNumber in pageContent.split():
            #     pass
            # else:
            #     labelforERRORs.configure(text="ERROR: National Insurance Number does not match report: "+P45s)
            #     x.close()
            #     os.remove(P45s[:-4]+'.jpg')
            #     Error = True
            #     break
            if Name in pageContent.split():
                pass
            else:
                labelforERRORs.configure(text="ERROR: Name does not match report: "+P45s+". Please run a new report and sort again")
                x.close()
                os.remove(P45s[:-4]+'.jpg')
                Error = True
                break
            x.close()
            os.remove(P45s[:-4]+'.jpg')
            PDFFile = PyPDF3.PdfFileReader(P45Folder+'/'+P45s)
            NumberOfPages = PDFFile.numPages
            Output_PDFFile = PyPDF3.PdfFileWriter()
            for i in range(NumberOfPages):
                Output_PDFFile.addPage(PDFFile.getPage(i))
            Output_PDFFile.encrypt(NatNumber)
            Output_PDFFile.write(open(P45Folder+'\\P45s Ready To Send\\'+PayrollNumber+" "+nameForRename+".pdf", 'wb'))
            FilesRemaned += 1
            newFileName = ''
            NatNumber = ""
            PayrollNumber = ""
            EmailAddress = ""
            Name = ""
            nameForRename = ''
            labetotalP45tosend.configure(text='File being Processed: '+str(FilesRemaned)+'/'+str(totalP45sInFolder))
            P45SenderWindow.update()
        P45stoSend = os.listdir(P45Folder+'\\P45s Ready To Send')
        for files4 in P45stoSend:
            if files4.endswith('.pdf') or files4.endswith('.PDF'):
                newFileName = ''
                EmailAddress = ""
                PayrollNumber = ''
                NameOnthePDFFile = ''
                files1 = ''
                files1 = Path(files4).stem 
                PayrollNumber = files1[:6]
                NameOnthePDFFile = files1[7:]
                newFileName = PayrollNumber+' '+NameOnthePDFFile
                EmailAddress = PaynumberToEmailDict.get(PayrollNumber)
                NatNumber =  POtoNatIndict.get(PayrollNumber)
                ListofP45stoEmailAddress.append(newFileName+'    ---->    '+EmailAddress)
                files1 = ''
                PayrollNumber = ''
                NameOnthePDFFile = ''
                newFileName = ''
                EmailAddress = ""  
        if Error == False:
            labetotalP45tosend.configure(text="Total ready to send is "+str(len(P45stoSend)-1))
            window = Toplevel(P45SenderWindow)
            window.title('List of P45s and Address for them to be sent to')
            window.geometry("900x700+500+200") 
            window.configure(bg="RoyalBlue4")
            window.attributes('-topmost',1)
            ListofP45toEmail = Listbox(window, bg="royalblue1",width=80, height=31, selectmode='single', font=('Times', 14))
            ListofP45toEmail.pack()
            ListofP45toEmail.insert(END,'Total P45s to be sent '+str(len(P45stoSend)-1))
            ListofP45toEmail.insert(END,'')
            for x in ListofP45stoEmailAddress:
                ListofP45toEmail.insert(END,x)
            Safety = 'Off'
            button_sort.configure(bg='green4')
            button_send.configure(bg='blue')
            P45SenderWindow.update()
        else:
            pass
        ListofP45stoEmailAddress = []
    def Send():
        global Safety
        global P45stoSend
        global SortSafety1
        global SortSafety2
        if Safety == 'Off':
            P45stoSend = os.listdir(P45Folder+'\\P45s Ready To Send')
            PayrollNumber = ""
            EmailAddress = ""
            NameOnthePDFFile = ""
            PayrollNumberCheck = ""
            Name = ""
            global FilesRemaned
            global P45SentCOunter
            global totalP45sInFolder
            for P45stos in P45stoSend:
                if P45stos.endswith('.pdf') or P45stos.endswith('.PDF'):
                    PayrollNumber = ""
                    EmailAddress = ""
                    NameOnthePDFFile = ""
                    PayrollNumberCheck = ""
                    Name = ""
                    NatNumber = ""
                    nameForRename = ""
                    file = Path(P45Folder+'\\P45s Ready To Send\\'+P45stos).stem           #### CHANGE FOLDER NAME HERE 
                    PayrollNumber = file[:6]
                    NameOnthePDFFile = file[7:]
                    EmailAddress = PaynumberToEmailDict.get(PayrollNumber)
                    PayrollNumberCheck = NamesList2ListtoPayNumber.get(NameOnthePDFFile)
                    if PayrollNumber == PayrollNumberCheck:
                        pass
                    else:
                        labelforERRORs.configure(text="Name of P45 does not match name on file when trying to send "+P45stos+". Please run a new report and sort again")
                        Safety = 'On'
                        break
                    outlook = win32.Dispatch('outlook.application')
                    mail = outlook.CreateItem(0)
                    mail.To = EmailAddress
                    mail.Subject = 'P45'
                    mail.Body = 'Please find your P45 enclosed in this email.\n\nThe password to open your P45, is the same password to open your wageslip. Please contract your old store manager if you are unsure on what this is, as we cannot give hints for the password over email.\n\nRegards\n\nTOFS'
                    mail.Attachments.Add(P45Folder+'\\P45s Ready To Send\\'+P45stos)
                    #mail.CC = "payroll@tofs.com"
                    mail.Send()
                    P45SentCOunter += 1
                    labetotalP45tosend.configure(text = "Amount of files sent: %d/%d" % (P45SentCOunter,totalP45sInFolder))
                    NatNumber = ""
                    PayrollNumber = ""
                    EmailAddress = ""
                    NameOnthePDFFile = ""
                    PayrollNumberCheck = ""
                    nameForRename = ""
                    Name = ""
                    shutil.move(P45Folder+'\\P45s Ready To Send\\'+P45stos,P45Folder+'\\P45s Ready To Send\\P45s Sent\\'+P45stos)
                    P45SenderWindow.update()
                    time.sleep(1)
            labetotalP45tosend.configure(text = "All emails sent %d/%d\n\nPlease remove all files in 'P45s To Send', ready for next process" % (P45SentCOunter,totalP45sInFolder))
            Safety = 'On'
            SortSafety1 = 'On'
            SortSafety2 = 'On'
            button_explore.configure(bg='red3')
            button_explore2.configure(bg='red3')
            button_sort.configure(bg='red3')
            button_send.configure(bg='red3')
            P45SenderWindow.update()
            P45SentCOunter = 0
        else:
            labelforERRORs.configure(text='Please sort the P45s before attempting to send')
    label_file_explorer = Label(P45SenderWindow,text = "Original Factory Shop Email Sender - By Nick",width = 100, height = 4,fg = "white",bg = 'RoyalBlue4', font=('Times', 13)) #RoyalBlue4
    button_explore = Button(P45SenderWindow,text = "Select Chris21 Leavers Report",bg = 'red3',width = 30, height = 2,command = browser, font=('Times', 15, 'bold'), fg = "yellow2")
    button_explore2 = Button(P45SenderWindow,text = "Select P45 Folder",bg = 'red3',width = 30, height = 2,command = browser2, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exit = Button(P45SenderWindow,text = "Exit",bg = 'snow4',width = 30,height = 2,command = sys.exit, font=('Times', 15, 'bold'), fg = "black")
    button_sort = Button(P45SenderWindow,text = "Sort",bg = 'red3',width = 30,height = 2,command = Sort, font=('Times', 15, 'bold'), fg = "yellow2")
    button_send = Button(P45SenderWindow,text = "Send",bg = 'red3',width = 30,height = 2,command = Send, font=('Times', 15, 'bold'), fg = "yellow2")
    labelfileopened = Label(P45SenderWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labetotalP45tosend = Label(P45SenderWindow,text = "",width = 50, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelforERRORs = Label(P45SenderWindow,text = "",width = 75, height = 2,fg = "red",bg = 'RoyalBlue4', font=('Times', 16))
    labetotalP45tosend = Label(P45SenderWindow,text = "",width = 75,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    LabelSpace1P45 = Label(P45SenderWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace2P45 = Label(P45SenderWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace3P45 = Label(P45SenderWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace4P45 = Label(P45SenderWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    TitleImage = os.getcwd()+'\\1519797862804.jpg'
    img = Image.open(TitleImage)
    img = img.resize((500, 100), Image.LANCZOS)
    img = ImageTk.PhotoImage(img)
    labelimage = Label(P45SenderWindow, image = img,width = 500, height = 100)
    def allpacks():
        labelimage.pack()
        label_file_explorer.pack()
        button_explore.pack()
        LabelSpace1P45.pack()
        button_explore2.pack()
        LabelSpace2P45.pack()
        button_sort.pack()
        LabelSpace3P45.pack()
        button_exit.pack()
        LabelSpace4P45.pack()
        button_send.pack()
        labelforERRORs.pack()
        labelfileopened.pack()
        labetotalP45tosend.pack()
        labetotalP45tosend.pack()
    allpacks()
    P45SenderWindow.mainloop()
def NESTSelected():
    NESTWindow=Toplevel(MainWindow)
    NESTWindow.title("Sending NEST emails")
    NESTWindow.geometry("900x800+500+100")
    NESTWindow.configure(bg="RoyalBlue4")
    NESTWindow.attributes('-topmost',1)
    nestlettersent = 0
    global NESTSendSafety1
    global NESTSendSafety2
    NESTSendSafety1 = 'On'
    NESTSendSafety2 = 'On'
    def NestExcelFile():
        global NESTExcelFIle
        global ExcelFileNEST
        global ExcelSheetNEST
        global AddressListNEST
        global TotalEmailAdressesNEST
        global NESTSendSafety1
        button_exploreExcelNEST.configure(bg='red3')
        button_exploreDOCxNEST.configure(bg='red3')
        button_sendNEST.configure(bg='red3')
        NESTSendSafety1 = 'On'
        labelERRORsfoRNEST.configure(text='')
        AddressListNEST = []
        NESTExcelFIle = filedialog.askopenfilename(parent=NESTWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select file",filetypes = (("Excel","*xlsx"),("all files","*.*")))
        NESTExcelFIle0 = Path(NESTExcelFIle).stem 
        labelfileopnedNEST.configure(text="File Opened: "+NESTExcelFIle0)
        ExcelFileNEST = openpyxl.load_workbook(NESTExcelFIle)
        ExcelSheetNEST = ExcelFileNEST.active
        for cell in ExcelSheetNEST['N']:
            if cell.value != None and '@' in cell.value:
                AddressListNEST.append(cell.value)
        TotalEmailAdressesNEST = len(AddressListNEST)
        if TotalEmailAdressesNEST > 0:
            pass
        else:
            labelERRORsfoRNEST.configure(text='ERROR: No email address found in Col "N"')
            NESTWindow.update()
            NESTSendSafety1 = 'On'
            return
        NESTWindo1 = Toplevel(NESTWindow)
        NESTWindo1.title('All PDF files in folder, please select another folder if wrong')
        NESTWindo1.geometry("900x700+500+200") 
        NESTWindo1.configure(bg="RoyalBlue4")
        NESTWindo1.attributes('-topmost',1)
        ListofEmailAddressListBoxNEST = Listbox(NESTWindo1, bg="royalblue1",width=80, height=31, selectmode='single', font=('Times', 14))
        ListofEmailAddressListBoxNEST.pack()
        ListofEmailAddressListBoxNEST.insert(END,'All Email Addressses in selected excel file, please choose another file if wrong')
        ListofEmailAddressListBoxNEST.insert(END,'')
        ListofEmailAddressListBoxNEST.insert(END,'Total emails to send to: '+str(TotalEmailAdressesNEST))
        ListofEmailAddressListBoxNEST.insert(END,'')
        for xNEST in AddressListNEST:
            ListofEmailAddressListBoxNEST.insert(END,xNEST)
        labelsentcounterNEST.configure(text="Total Email Addresses to send to: "+str(TotalEmailAdressesNEST))
        NESTSendSafety1 = 'Off'
        button_exploreExcelNEST.configure(bg='green4')
    def SelectNESTDocx():
        global NESTSendSafety2
        global enrollmentletterNEST
        button_exploreDOCxNEST.configure(bg='red3')
        button_sendNEST.configure(bg='red3')
        NESTSendSafety2 = 'On'
        enrollmentletterNEST = ''
        labelERRORsfoRNEST.configure(text='')
        if NESTSendSafety1 == 'On':
            labelERRORsfoRNEST.configure(text='Please select the excel report first')
            button_exploreDOCxNEST.configure(bg='red3')
            return
        try:
            enrollmentletterNEST = filedialog.askopenfilename(parent=NESTWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select file",filetypes = (("DOCx","*docx"),))
            filename0NEST = Path(enrollmentletterNEST).stem 
            labelfileopnedDocxNEST.configure(text="File To Be Sent: "+filename0NEST)
        except:
            labelERRORsfoRNEST.configure(text='Please select a valid DOCx file')
            return
        if len(enrollmentletterNEST) == 0:
            NESTSendSafety2 = 'On'
            button_exploreDOCxNEST.configure(bg='red3')
            labelERRORsfoRNEST.configure(text='Please select a valid DOCx file')
            return
        else:
            NESTSendSafety2 = 'Off'
            button_exploreDOCxNEST.configure(bg='green4')
            button_sendNEST.configure(bg='blue')
    def SendNEST():
        global NESTSendSafety1
        global NESTSendSafety2
        if NESTSendSafety1 == 'On':
            labelERRORsfoRNEST.configure(text='Please select Excel file before sending')
            NESTWindow.update()
            return
        if NESTSendSafety2 == 'On':
            labelERRORsfoRNEST.configure(text='Please select DOCx before sending')
            NESTWindow.update()
            return
        else:         
            global TotalEmailAdressesNEST
            global nestlettersent
            global enrollmentletterNEST
            nestlettersent = 0
            for NESTEmails in AddressListNEST:
                outlook = win32.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)
                mail.To = NESTEmails
                mail.Subject = 'NEST ENROLLMENT LETTER'
                mail.Body = 'Hi\n\nEnclosed is your NEST enrollment letter for the company pension\n\nKind Regards,\n\nPAYROLL\n\nTOFS'
                #mail.CC = "payroll@tofs.com"
                mail.Attachments.Add(enrollmentletterNEST)
                mail.Send()
                nestlettersent += 1
                labelsentcounterNEST.configure(text = "Amount of files sent: %d/%d" % (nestlettersent,TotalEmailAdressesNEST))
                time.sleep(1)
                NESTWindow.update()
            labelsentcounterNEST.configure(text = "All emails sent %d/%d" % (nestlettersent,TotalEmailAdressesNEST))
            NESTSendSafety1 = 'On'
            NESTSendSafety2 = 'On'
            button_exploreExcelNEST.configure(bg='red3')
            button_exploreDOCxNEST.configure(bg='red3')
            button_sendNEST.configure(bg='red3')
            NESTWindow.update()
    label_file_explorerNEST = Label(NESTWindow,text = "Original Factory Shop Email Sender - By Nick",width = 100, height = 4,fg = "white",bg = 'RoyalBlue4', font=('Times', 13))
    button_exploreExcelNEST = Button(NESTWindow,text = "Select Excel File for Sending",bg = 'red3',width = 30, height = 2,command = NestExcelFile, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exploreDOCxNEST = Button(NESTWindow,text = "Select NEST template file to be Sent",bg = 'red3',width = 30, height = 2,command = SelectNESTDocx, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exitNEST = Button(NESTWindow,text = "Exit",bg = 'snow4',width = 30,height = 2,command = sys.exit, font=('Times', 15, 'bold'), fg = "black")
    button_sendNEST = Button(NESTWindow,text = "Send",bg = 'red3',width = 30,height = 2,command = SendNEST, font=('Times', 15, 'bold'), fg = "yellow2")
    labelERRORsfoRNEST = Label(NESTWindow,text = "",width = 75, height = 2,fg = "red",bg = 'RoyalBlue4', font=('Times', 16))
    labelfileopnedNEST = Label(NESTWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelsentcounterNEST = Label(NESTWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelfileopnedDocxNEST = Label(NESTWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    LabelSpace1NEST = Label(NESTWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace2NEST = Label(NESTWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace3NEST = Label(NESTWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    TitleImageNEST = os.getcwd()+'\\1519797862804.jpg'
    imgNEST = Image.open(TitleImageNEST)
    imgNEST = imgNEST.resize((500, 100), Image.LANCZOS)
    imgNEST = ImageTk.PhotoImage(imgNEST)
    labelimageNEST = Label(NESTWindow, image = imgNEST,width = 500, height = 100)
    def allpacks():
        labelimageNEST.pack()
        label_file_explorerNEST.pack()
        button_exploreExcelNEST.pack()
        LabelSpace1NEST.pack()
        button_exploreDOCxNEST.pack()
        LabelSpace2NEST.pack()
        button_exitNEST.pack()
        LabelSpace3NEST.pack()
        button_sendNEST.pack()
        labelERRORsfoRNEST.pack()
        labelfileopnedNEST.pack()
        labelfileopnedDocxNEST.pack()
        labelsentcounterNEST.pack()
    allpacks()
    NESTWindow.mainloop()   
TitleImage = os.getcwd()+'\\1519797862804.jpg'  
img = Image.open(TitleImage)
img = img.resize((500, 100), Image.LANCZOS)
img = ImageTk.PhotoImage(img)
labelimage = Label(MainWindow, image = img,width = 500, height = 100)
Labelfill = Label(MainWindow, width = 30, height = 3,bg = 'RoyalBlue4')
Labelfill1 = Label(MainWindow, width = 30, height = 3,bg = 'RoyalBlue4')
P45WindowButton = Button(MainWindow,text = "Send P45s",bg = 'RoyalBlue1',width = 30,height = 2, font=('Times', 15, 'bold'), fg = "yellow2",command = P45Selected) #command = Sort
NESTWindowButton = Button(MainWindow,text = "Send NEST files",bg = 'RoyalBlue1',width = 30,height = 2, font=('Times', 15, 'bold'), fg = "yellow2",command = NESTSelected) #command = P45Selected
labelimage.pack()
Labelfill.pack()
P45WindowButton.pack()
Labelfill1.pack()
NESTWindowButton.pack()
MainWindow.mainloop()


