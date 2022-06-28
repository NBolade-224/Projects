###### FILE EMAIL SENDER
## TO DO 
## MAKE SO THAT ALL THE FOLDERS IN THE TWO SEND FILE GET SENT(WHETHER PROCESSED OR NOT)
## ADD SHUTIL MOVE SO FILES THAT ARE MOVED TO THE SENDING FOLDER ARE REMOVED
## ADD ANOTHER SHUTIL MOVE SO THAT FILES SENT ARE ALSO MOVED TO THAT FOLDER
## DO MORE TTESTING!!!
## COMPILE AND MAKE IT WORKS ON INDEPENDENT LAPTOP (IT FAILED TO SORT ON KARENS, FIGURE OUT WHY...)
## COMPILE IT ON WORKDESKTOP AND TRY IN HOMEPC
### need to set automatic creation of ready to send folder and sent folder. 
from tkinter import *
from tkinter import filedialog
from pathlib import Path
import win32com.client as win32
import pytesseract, os, PyPDF3, openpyxl, sys, shutil
from openpyxl.utils.cell import get_column_letter
from pdf2image import convert_from_path
from PIL import ImageTk, Image
import time
MainWindow=Tk()
MainWindow.title("Sending Emails")
MainWindow.geometry("600x500+600+200")
MainWindow.configure(bg="RoyalBlue4")
def FILESelected():
    PopplerPath = os.getcwd()+'\\poppler-0.68.0_x86\\poppler-0.68.0\\bin'
    Tesspath = os.getcwd()+'\\Tesseract-OCR\\tesseract.exe'
    poppler_path = PopplerPath
    pytesseract.pytesseract.tesseract_cmd = Tesspath
    FILESenderWindow=Toplevel(MainWindow)
    FILESenderWindow.title("Sending Emails")
    FILESenderWindow.geometry("900x800+500+100")
    FILESenderWindow.configure(bg="RoyalBlue4")
    FILESenderWindow.attributes('-topmost',1)
    global Safety
    global SortSafety1
    global SortSafety2
    global SortSafety3
    global FILEsInFolderList
    global FILTESToSend
    global FILESentCOunter
    global FilesRemaned
    FILEsInFolderList = []
    FILTESToSend = []
    FILESentCOunter = 0
    FilesRemaned = 0
    Safety = 'On'
    SortSafety1 = 'On'
    SortSafety2 = 'On'
    SortSafety3 = 'Off'
    def browser():
        labelforERRORs.configure(text='')
        global filename
        global ExcelFile
        global ExcelSheet
        global EmailAddressList
        global PRNumberList
        global NamesList
        global PaynumberToEmailDict
        global EmailtoNameDict
        global NatInNumberList
        global POtoNatIndict
        global NamesList2List
        global Name1toName2Dict
        global SortSafety1
        global NamesList2ListtoPayNumber
        NamesList2ListtoPayNumber = []
        EmailAddressList = []
        PRNumberList = []
        NamesList = []
        NatInNumberList = []
        NamesList2List = []
        try:
            filename = filedialog.askopenfilename(parent=FILESenderWindow, initialdir = "C:\\Users\\folder\\Desktop",title = "Select file",filetypes = (("Excel","*xlsx"),))
            ExcelFile = openpyxl.load_workbook(filename)
            ExcelSheet = ExcelFile.active
            filename0 = Path(filename).stem 
            labelfileopened.configure(text="File Opened: "+filename0)
        except:
            labelforERRORs.configure(text='Please select a valid excel file')
            SortSafety1 = 'On'
            return
        for cell in ExcelSheet['I']:
            EmailAddressList.append(cell.value)
        for cell in ExcelSheet['C']:
            PRNumberList.append(cell.value)
        for cell in ExcelSheet['D']:
            NatInNumberList.append(cell.value)    
        for cell in ExcelSheet['B']:
            NamesList.append(cell.value)
        for cell in ExcelSheet['E']:
            NamesList2List.append(cell.value)
        PaynumberToEmailDict = dict(zip(PRNumberList, EmailAddressList))
        EmailtoNameDict = dict(zip(EmailAddressList, NamesList))
        Name1toName2Dict = dict(zip(NamesList, NamesList2List))
        POtoNatIndict = dict(zip(PRNumberList, NatInNumberList))
        NamesList2ListtoPayNumber = dict(zip(NamesList2List, PRNumberList))
        SortSafety1 = 'Off'
        FILESenderWindow.update()
    def browser2():
        global SortSafety2
        global SortSafety3
        labelforERRORs.configure(text="")
        global FILEsInFolderList
        global totalFILEsInFolder
        global FILEFolder
        FILEsInFolderList = []
        FILEFolder = filedialog.askdirectory(parent=FILESenderWindow, initialdir = "C:\\Users\\folder\\Desktop",title = "Select FILE Folder")
        try:
            FILEsInFolder = os.listdir(FILEFolder)
        except:
            labelforERRORs.configure(text="ERROR: No folder selected")
            return
        SortSafety3 = 'Off'
        if 'FILEs Ready To Send' in FILEFolder[-18:] or 'FILEs Sent' in FILEFolder[-9:]:
            labelforERRORs.configure(text="ERROR: Please do not select the 'FILEs To Send' folder or subfolders")
            FILESenderWindow.update()
            return
        for files in FILEsInFolder:
            if files.endswith('.pdf') or files.endswith('.PDF'):
                FILEsInFolderList.append(files)
        totalFILEsInFolder = len(FILEsInFolderList)
        if totalFILEsInFolder == 0:
            labelforERRORs.configure(text="ERROR: No PDF files found in folder, please select correct folder")
            FILESenderWindow.update()
            return
        window1 = Toplevel(FILESenderWindow)
        window1.title('All PDF files in folder, please select another folder if wrong')
        window1.geometry("900x700+500+200") 
        window1.configure(bg="RoyalBlue4")
        window1.attributes('-topmost',1)
        ListofPDFfilesinFolder = Listbox(window1, bg="grey",width=80, height=31, selectmode='single', font=('Times', 14))
        ListofPDFfilesinFolder.pack()
        ListofPDFfilesinFolder.insert(END,'All PDF files in selected folder, please choose another folder if wrong')
        ListofPDFfilesinFolder.insert(END,'')
        ListofPDFfilesinFolder.insert(END,'Total files in folder: '+str(totalFILEsInFolder))
        ListofPDFfilesinFolder.insert(END,'')
        for x in FILEsInFolderList:
            ListofPDFfilesinFolder.insert(END,x)
        labetotalFILEtosend.configure(text="FILEs to be processed: "+str(totalFILEsInFolder))
        try:
            if len(os.listdir(FILEFolder+'\\FILEs Ready To Send')) > 1:
                labelforERRORs.configure(text="Warning, files found in 'FILEs ToSend' folder, please delete these files before sorting")
                SortSafety3 = 'On'
                FILESenderWindow.update()
                return
            else:
                SortSafety2 = 'Off'
                FILESenderWindow.update()
        except:
            SortSafety2 = 'Off'
            FILESenderWindow.update()
            pass
    def Sort():
        if  SortSafety3 == 'On':
            labelforERRORs.configure(text="Warning, files found in 'FILEs ToSend' folder, please delete these files before sorting")
            FILESenderWindow.update()
            return
        if SortSafety1 == 'On':
            labelforERRORs.configure(text="Please select a leavers report before sorting")
            FILESenderWindow.update()
            return
        if SortSafety2 == 'On':
            labelforERRORs.configure(text="Please select the FILE folder before sorting")
            FILESenderWindow.update()
            return
        ListofFILTESToEmailAddress = []
        Path(FILEFolder+'\\FILEs Ready To Send\\FILEs Sent').mkdir(parents=True, exist_ok=True)
        Path(FILEFolder+'\\FILEs Ready To Send').mkdir(parents=True, exist_ok=True)
        labelforERRORs.configure(text="")
        PRNumber = ""
        EmailAddress = "" 
        Name = ""
        labetotalFILEtosend.configure(text="")
        global FILEsInFolderList
        global FilesRemaned
        global FILTESToSend
        global Safety
        global totalFILEsInFolder
        Safety = 'On'
        FILTESToSend = []
        FilesRemaned = 0
        Error = False
        for FILEs in FILEsInFolderList:
            PRNumber = ""
            EmailAddress = ""
            Name = ""
            NatNumber = ""
            newFileName = ''
            nameForRename = ''
            page = convert_from_path(FILEFolder+"\\"+FILEs, 600,poppler_path = PopplerPath) 
            page[0].save(FILEs[:-4]+'.jpg', 'JPEG')
            x = Image.open(FILEs[:-4]+'.jpg')
            pageContent = pytesseract.image_to_string(x)
            for content in pageContent.split():
                if len(content) == 6 and content.startswith("0"): ## and is int
                    PRNumber = content
            if PRNumber == "":
                labelforERRORs.configure(text="ERROR: Failed to find PR number in file: "+FILEs+". Please run a new report and sort again")
                Error = True
                x.close()
                os.remove(FILEs[:-4]+'.jpg')
                break
            EmailAddress = PaynumberToEmailDict.get(PRNumber)
            Name = EmailtoNameDict.get(EmailAddress)
            nameForRename = Name1toName2Dict.get(Name)
            NatNumber =  POtoNatIndict.get(PRNumber)
            if Name in pageContent.split():
                pass
            else:
                labelforERRORs.configure(text="ERROR: Name does not match report: "+FILEs+". Please run a new report and sort again")
                x.close()
                os.remove(FILEs[:-4]+'.jpg')
                Error = True
                break
            x.close()
            os.remove(FILEs[:-4]+'.jpg')
            PDFFile = PyPDF3.PdfFileReader(FILEFolder+'/'+FILEs)
            NumberOfPages = PDFFile.numPages
            Output_PDFFile = PyPDF3.PdfFileWriter()
            for i in range(NumberOfPages):
                Output_PDFFile.addPage(PDFFile.getPage(i))
            Output_PDFFile.encrypt(NatNumber)
            Output_PDFFile.write(open(FILEFolder+'\\FILEs Ready To Send\\'+PRNumber+" "+nameForRename+".pdf", 'wb'))
            FilesRemaned += 1
            newFileName = ''
            NatNumber = ""
            PRNumber = ""
            EmailAddress = ""
            Name = ""
            nameForRename = ''
            labetotalFILEtosend.configure(text='File being Processed: '+str(FilesRemaned)+'/'+str(totalFILEsInFolder))
            FILESenderWindow.update()
        FILTESToSend = os.listdir(FILEFolder+'\\FILEs Ready To Send')
        for files4 in FILTESToSend:
            if files4.endswith('.pdf') or files4.endswith('.PDF'):
                newFileName = ''
                EmailAddress = ""
                PRNumber = ''
                NameOnthePDFFile = ''
                files1 = ''
                files1 = Path(files4).stem 
                PRNumber = files1[:6]
                NameOnthePDFFile = files1[7:]
                newFileName = PRNumber+' '+NameOnthePDFFile
                EmailAddress = PaynumberToEmailDict.get(PRNumber)
                NatNumber =  POtoNatIndict.get(PRNumber)
                ListofFILTESToEmailAddress.append(newFileName+'    ---->    '+EmailAddress)
                files1 = ''
                PRNumber = ''
                NameOnthePDFFile = ''
                newFileName = ''
                EmailAddress = ""  
        if Error == False:
            labetotalFILEtosend.configure(text="Total ready to send is "+str(len(FILTESToSend)-1))
            window = Toplevel(FILESenderWindow)
            window.title('List of FILEs and Address for them to be sent to')
            window.geometry("900x700+500+200") 
            window.configure(bg="RoyalBlue4")
            window.attributes('-topmost',1)
            ListofFILEtoEmail = Listbox(window, bg="grey",width=80, height=31, selectmode='single', font=('Times', 14))
            ListofFILEtoEmail.pack()
            ListofFILEtoEmail.insert(END,'Total FILEs to be sent '+str(len(FILTESToSend)-1))
            ListofFILEtoEmail.insert(END,'')
            for x in ListofFILTESToEmailAddress:
                ListofFILEtoEmail.insert(END,x)
            Safety = 'Off'
            FILESenderWindow.update()
        else:
            pass
        ListofFILTESToEmailAddress = []
    def Send():
        global Safety
        global FILTESToSend
        if Safety == 'Off':
            FILTESToSend = os.listdir(FILEFolder+'\\FILEs Ready To Send')
            PRNumber = ""
            EmailAddress = ""
            NameOnthePDFFile = ""
            PRNumberCheck = ""
            Name = ""
            global FilesRemaned
            global FILESentCOunter
            global totalFILEsInFolder
            for FILTESTos in FILTESToSend:
                if FILTESTos.endswith('.pdf') or FILTESTos.endswith('.PDF'):
                    PRNumber = ""
                    EmailAddress = ""
                    NameOnthePDFFile = ""
                    PRNumberCheck = ""
                    Name = ""
                    NatNumber = ""
                    nameForRename = ""
                    file = Path(FILEFolder+'\\FILEs Ready To Send\\'+FILTESTos).stem           #### CHANGE FOLDER NAME HERE 
                    PRNumber = file[:6]
                    NameOnthePDFFile = file[7:]
                    EmailAddress = PaynumberToEmailDict.get(PRNumber)
                    PRNumberCheck = NamesList2ListtoPayNumber.get(NameOnthePDFFile)
                    if PRNumber == PRNumberCheck:
                        pass
                    else:
                        labelforERRORs.configure(text="Name of FILE does not match name on file when trying to send "+FILTESTos+". Please run a new report and sort again")
                        Safety = 'On'
                        break
                    outlook = win32.Dispatch('outlook.application')
                    mail = outlook.CreateItem(0)
                    mail.To = EmailAddress
                    mail.Subject = 'FILE'
                    mail.Body = "Messager Here"
                    mail.Attachments.Add(FILEFolder+'\\FILEs Ready To Send\\'+FILTESTos)
                    mail.CC = "addresss"
                    mail.Send()
                    FILESentCOunter += 1
                    labetotalFILEtosend.configure(text = "Amount of files sent: %d/%d" % (FILESentCOunter,totalFILEsInFolder))
                    NatNumber = ""
                    PRNumber = ""
                    EmailAddress = ""
                    NameOnthePDFFile = ""
                    PRNumberCheck = ""
                    nameForRename = ""
                    Name = ""
                    shutil.move(FILEFolder+'\\FILEs Ready To Send\\'+FILTESTos,FILEFolder+'\\FILEs Ready To Send\\FILEs Sent\\'+FILTESTos)
                    FILESenderWindow.update()
                    time.sleep(1)
            labetotalFILEtosend.configure(text = "All emails sent %d/%d\n\nPlease remove all files in 'FILEs To Send', ready for next process" % (FILESentCOunter,totalFILEsInFolder))
            FILESenderWindow.update()
            FILESentCOunter = 0
        else:
            labelforERRORs.configure(text='Please sort the FILEs before attempting to send')
    label_file_explorer = Label(FILESenderWindow,text = "",width = 100, height = 4,fg = "white",bg = 'RoyalBlue4', font=('Times', 13)) #RoyalBlue4
    button_explore = Button(FILESenderWindow,text = "Select Report",bg = 'RoyalBlue4',width = 30, height = 2,command = browser, font=('Times', 15, 'bold'), fg = "yellow2")
    button_explore2 = Button(FILESenderWindow,text = "Select FILE Folder",bg = 'RoyalBlue4',width = 30, height = 2,command = browser2, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exit = Button(FILESenderWindow,text = "Exit",bg = 'RoyalBlue4',width = 30,height = 2,command = sys.exit, font=('Times', 15, 'bold'), fg = "yellow2")
    button_sort = Button(FILESenderWindow,text = "Sort",bg = 'RoyalBlue4',width = 30,height = 2,command = Sort, font=('Times', 15, 'bold'), fg = "yellow2")
    button_send = Button(FILESenderWindow,text = "Send",bg = 'RoyalBlue4',width = 30,height = 2,command = Send, font=('Times', 15, 'bold'), fg = "yellow2")
    labelfileopened = Label(FILESenderWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labetotalFILEtosend = Label(FILESenderWindow,text = "",width = 50, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelforERRORs = Label(FILESenderWindow,text = "",width = 75, height = 2,fg = "red",bg = 'RoyalBlue4', font=('Times', 16))
    labetotalFILEtosend = Label(FILESenderWindow,text = "",width = 75,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    TitleImage = os.getcwd()+'\\1519797862804.jpg'
    img = Image.open(TitleImage)
    img = img.resize((500, 100), Image.LANCZOS)
    img = ImageTk.PhotoImage(img)
    labelimage = Label(FILESenderWindow, image = img,width = 500, height = 100)
    def allpacks():
        labelimage.pack()
        label_file_explorer.pack()
        button_explore.pack()
        button_explore2.pack()
        button_sort.pack()
        button_exit.pack()
        button_send.pack()
        labelforERRORs.pack()
        labelfileopened.pack()
        labetotalFILEtosend.pack()
        labetotalFILEtosend.pack()
    allpacks()
    FILESenderWindow.mainloop()
def NTESTSelected():
    NTESTWindow=Toplevel(MainWindow)
    NTESTWindow.title("Sending NTEST emails")
    NTESTWindow.geometry("900x800+500+100")
    NTESTWindow.configure(bg="RoyalBlue4")
    NTESTWindow.attributes('-topmost',1)
    nTESTlettersent = 0
    global NTESTSendSafety1
    global NTESTSendSafety2
    NTESTSendSafety1 = 'On'
    NTESTSendSafety2 = 'On'
    def NTESTExcelFile():
        global NTESTExcelFIle
        global ExcelFileNTEST
        global ExcelSheetNTEST
        global AddressListNTEST
        global TotalEmailAdressesNTEST
        global NTESTSendSafety1
        NTESTSendSafety1 = 'On'
        labelERRORsfoRNTEST.configure(text='')
        AddressListNTEST = []
        NTESTExcelFIle = filedialog.askopenfilename(parent=NTESTWindow, initialdir = "C:\\Users\\folder\\Desktop",title = "Select file",filetypes = (("Excel","*xlsx"),("all files","*.*")))
        labelfileopnedNTEST.configure(text="File Opened: "+NTESTExcelFIle)
        ExcelFileNTEST = openpyxl.load_workbook(NTESTExcelFIle)
        ExcelSheetNTEST = ExcelFileNTEST.active
        for cell in ExcelSheetNTEST['N']:
            if cell.value != None and '@' in cell.value:
                AddressListNTEST.append(cell.value)
        TotalEmailAdressesNTEST = len(AddressListNTEST)
        if TotalEmailAdressesNTEST > 0:
            pass
        else:
            labelERRORsfoRNTEST.configure(text='ERROR: No email address found in Col "N"')
            NTESTWindow.update()
            NTESTSendSafety1 = 'On'
            return
        NTESTWindo1 = Toplevel(NTESTWindow)
        NTESTWindo1.title('All PDF files in folder, please select another folder if wrong')
        NTESTWindo1.geometry("900x700+500+200") 
        NTESTWindo1.configure(bg="RoyalBlue4")
        NTESTWindo1.attributes('-topmost',1)
        ListofEmailAddressListBoxNTEST = Listbox(NTESTWindo1, bg="grey",width=80, height=31, selectmode='single', font=('Times', 14))
        ListofEmailAddressListBoxNTEST.pack()
        ListofEmailAddressListBoxNTEST.insert(END,'All Email Addressses in selected excel file, please choose another file if wrong')
        ListofEmailAddressListBoxNTEST.insert(END,'')
        ListofEmailAddressListBoxNTEST.insert(END,'Total emails to send to: '+str(TotalEmailAdressesNTEST))
        ListofEmailAddressListBoxNTEST.insert(END,'')
        for xNTEST in AddressListNTEST:
            ListofEmailAddressListBoxNTEST.insert(END,xNTEST)
        labelsentcounterNTEST.configure(text="Total Email Addresses to send to: "+str(TotalEmailAdressesNTEST))
        NTESTSendSafety1 = 'Off'
    def SelectNTESTDocx():
        global NTESTSendSafety2
        global enrollmentletterNTEST
        NTESTSendSafety2 = 'On'
        enrollmentletterNTEST = ''
        labelERRORsfoRNTEST.configure(text='')
        try:
            enrollmentletterNTEST = filedialog.askopenfilename(parent=NTESTWindow, initialdir = "C:\\Users\\folder\\Desktop",title = "Select file",filetypes = (("DOCx","*docx"),))
            filename0NTEST = Path(enrollmentletterNTEST).stem 
            labelfileopnedDocxNTEST.configure(text="File To Be Sent: "+filename0NTEST)
            NTESTSendSafety2 = 'Off'
        except:
            labelERRORsfoRNTEST.configure(text='Please select a valid DOCx file')
            return
    def SendNTEST():
        global NTESTSendSafety1
        global NTESTSendSafety2
        if NTESTSendSafety1 == 'On':
            labelERRORsfoRNTEST.configure(text='Please select Excel file before sending')
            NTESTWindow.update()
            return
        if NTESTSendSafety2 == 'On':
            labelERRORsfoRNTEST.configure(text='Please select DOCx before sending')
            NTESTWindow.update()
            return
        else:         
            global TotalEmailAdressesNTEST
            global nTESTlettersent
            global enrollmentletterNTEST
            nTESTlettersent = 0
            for NTESTEmails in AddressListNTEST:
                outlook = win32.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)
                mail.To = NTESTEmails
                mail.Subject = 'NTEST ENROLLMENT LETTER'
                mail.Body = 'Message here'
                mail.CC = "Address"
                mail.Attachments.Add(enrollmentletterNTEST)
                mail.Send()
                nTESTlettersent += 1
                labelsentcounterNTEST.configure(text = "Amount of files sent: %d/%d" % (nTESTlettersent,TotalEmailAdressesNTEST))
                time.sleep(1)
                NTESTWindow.update()
            labelsentcounterNTEST.configure(text = "All emails sent %d/%d" % (nTESTlettersent,TotalEmailAdressesNTEST))
            NTESTWindow.update()
    label_file_explorerNTEST = Label(NTESTWindow,text = "Original Factory Shop Email Sender - By Nick",width = 100, height = 4,fg = "white",bg = 'RoyalBlue4', font=('Times', 13))
    button_exploreExcelNTEST = Button(NTESTWindow,text = "Select Excel File for Sending",bg = 'RoyalBlue4',width = 30, height = 2,command = NTESTExcelFile, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exploreDOCxNTEST = Button(NTESTWindow,text = "Select NTEST template file to be Sent",bg = 'RoyalBlue4',width = 30, height = 2,command = SelectNTESTDocx, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exitNTEST = Button(NTESTWindow,text = "Exit",bg = 'RoyalBlue4',width = 30,height = 2,command = sys.exit, font=('Times', 15, 'bold'), fg = "yellow2")
    button_sendNTEST = Button(NTESTWindow,text = "Send",bg = 'RoyalBlue4',width = 30,height = 2,command = SendNTEST, font=('Times', 15, 'bold'), fg = "yellow2")
    labelERRORsfoRNTEST = Label(NTESTWindow,text = "",width = 75, height = 2,fg = "red",bg = 'RoyalBlue4', font=('Times', 16))
    labelfileopnedNTEST = Label(NTESTWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelsentcounterNTEST = Label(NTESTWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelfileopnedDocxNTEST = Label(NTESTWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    TitleImageNTEST = os.getcwd()+'\\1519797862804.jpg'
    imgNTEST = Image.open(TitleImageNTEST)
    imgNTEST = imgNTEST.resize((500, 100), Image.LANCZOS)
    imgNTEST = ImageTk.PhotoImage(imgNTEST)
    labelimageNTEST = Label(NTESTWindow, image = imgNTEST,width = 500, height = 100)
    def allpacks():
        labelimageNTEST.pack()
        label_file_explorerNTEST.pack()
        button_exploreExcelNTEST.pack()
        button_exploreDOCxNTEST.pack()
        button_exitNTEST.pack()
        button_sendNTEST.pack()
        labelERRORsfoRNTEST.pack()
        labelfileopnedNTEST.pack()
        labelfileopnedDocxNTEST.pack()
        labelsentcounterNTEST.pack()
    allpacks()
    NTESTWindow.mainloop()   
TitleImage = os.getcwd()+'\\1519797862804.jpg'  
img = Image.open(TitleImage)
img = img.resize((500, 100), Image.LANCZOS)
img = ImageTk.PhotoImage(img)
labelimage = Label(MainWindow, image = img,width = 500, height = 100)
Labelfill = Label(MainWindow, width = 30, height = 3,bg = 'RoyalBlue4')
Labelfill1 = Label(MainWindow, width = 30, height = 3,bg = 'RoyalBlue4')
FILEWindowButton = Button(MainWindow,text = "Send FILEs",bg = 'RoyalBlue4',width = 30,height = 2, font=('Times', 15, 'bold'), fg = "yellow2",command = FILESelected) #command = Sort
NTESTWindowButton = Button(MainWindow,text = "Send NTEST files",bg = 'RoyalBlue4',width = 30,height = 2, font=('Times', 15, 'bold'), fg = "yellow2",command = NTESTSelected) #command = FILESelected
labelimage.pack()
Labelfill.pack()
FILEWindowButton.pack()
Labelfill1.pack()
NTESTWindowButton.pack()
MainWindow.mainloop()


