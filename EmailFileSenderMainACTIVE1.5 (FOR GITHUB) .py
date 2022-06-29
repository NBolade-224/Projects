
from tkinter import *
from tkinter import filedialog
from pathlib import Path
import win32com.client as win32
import pytesseract, os, PyPDF3, openpyxl, sys, shutil
from openpyxl.utils.cell import get_column_letter
from pdf2image import convert_from_path
from PIL import ImageTk, Image
import time
MainWindow=Tk()
MainWindow.title("Sending Emails")
MainWindow.geometry("600x500+600+200")
MainWindow.configure(bg="RoyalBlue4")
def FiLeSelected():
    PopplerPath = os.getcwd()+'\\poppler-0.68.0_x86\\poppler-0.68.0\\bin'
    Tesspath = os.getcwd()+'\\Tesseract-OCR\\tesseract.exe'
    pytesseract.pytesseract.tesseract_cmd = Tesspath
    FiLeSenderWindow=Toplevel(MainWindow)
    FiLeSenderWindow.title("Sending Emails")
    FiLeSenderWindow.geometry("900x800+500+100")
    FiLeSenderWindow.configure(bg="RoyalBlue4")
    FiLeSenderWindow.attributes('-topmost',1)
    global SortSafety1
    global SortSafety2
    SortSafety1 = True
    SortSafety2 = True
    def browser():
        labelforERRORs.configure(text='')
        button_explore.configure(bg='red3')
        #################
        ### GLOBALS ####
        ##################
        global SortSafety1
        global RefNumberrToEmailDict
        global RefNumberrtoNameDict1
        global RefNumberrtoNameDict2
        global RefNumberrtoNameIntialsDict
        global RefNumberrtoNATDict
        #########################
        ### LIST FOR DICTS#########
        #########################
        EmailAddressList = []
        RefNumberList = []
        NamesList1 = []
        NamesList2 = []
        NatInNumberList = []
        NameInitialsList = []
        AddressListFiLe = []   ### To check that column has emails present
        try:
            filename = filedialog.askopenfilename(parent=FiLeSenderWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select file",filetypes = (("Excel","*xlsx"),))
            ExcelFile = openpyxl.load_workbook(filename)
            ExcelSheet = ExcelFile.active
            filename0 = Path(filename).stem 
            labelfileopened.configure(text="File Opened: "+filename0)
        except:
            labelforERRORs.configure(text='Please select a valid excel file')
            SortSafety1 = True
            return
        for cell in ExcelSheet['I']:
            if cell.value != None and '@' in cell.value:
                AddressListFiLe.append(cell.value)
        TotalEmailAdressesFiLe = len(AddressListFiLe)
        if TotalEmailAdressesFiLe > 0:
            AddressListFiLe = []
            pass
        else:
            labelforERRORs.configure(text='ERROR: No email address found in Col "I"\nPlease select Report')
            FiLeSenderWindow.update()
            SortSafety1 = True
            return
        for cell in ExcelSheet['I']:
            EmailAddressList.append(cell.value)
        for cell in ExcelSheet['C']:
            RefNumberList.append(cell.value)
        for cell in ExcelSheet['D']:
            NatInNumberList.append(cell.value)    
        for cell in ExcelSheet['A']:
            NamesList1.append(cell.value)
        for cell in ExcelSheet['B']:
            NamesList2.append(cell.value)
        for cell in ExcelSheet['E']:
            NameInitialsList.append(cell.value)
        RefNumberrToEmailDict = dict(zip(RefNumberList, EmailAddressList))
        RefNumberrtoNameDict1 = dict(zip(RefNumberList, NamesList1))
        RefNumberrtoNameDict2 = dict(zip(RefNumberList, NamesList2))
        RefNumberrtoNameIntialsDict = dict(zip(RefNumberList, NameInitialsList))
        RefNumberrtoNATDict = dict(zip(RefNumberList, NatInNumberList))
        button_explore.configure(bg='green4')
        SortSafety1 = False
        FiLeSenderWindow.update()
    def browser2():
        global SortSafety1
        global SortSafety2
        global FiLesInToSortFolder   ### Files that end with .PDF
        global TotalFiLesToBeSorted  ### NEED TO FIND TOTAL FILES TO BE SENT? E
        global FiLesFolderToSort    ### All files in selected folder, needed global to refence location for subfolders
        labelforERRORs.configure(text="")
        button_explore2.configure(bg='red3')
        SortSafety2 = True
        if SortSafety1 == True:
            labelforERRORs.configure(text="Please select a report first")
            return
        FiLesInToSortFolder = []
        FiLesFolderToSort = filedialog.askdirectory(parent=FiLeSenderWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select FiLe Folder")
        try:
            os.listdir(FiLesFolderToSort)
        except:
            labelforERRORs.configure(text="ERROR: No folder selected")
            return
        if 'FiLes Ready To Send' in FiLesFolderToSort[-18:] or 'FiLes Sent' in FiLesFolderToSort[-9:]:
            labelforERRORs.configure(text="ERROR: Please do not select the 'FiLes To Send' folder or subfolders")
            FiLeSenderWindow.update()
            return
        for files in os.listdir(FiLesFolderToSort):
            if files.endswith('.pdf') or files.endswith('.PDF'):
                FiLesInToSortFolder.append(files)
        TotalFiLesToBeSorted = len(FiLesInToSortFolder)
        if TotalFiLesToBeSorted == 0:
            labelforERRORs.configure(text="ERROR: No PDF files found in folder, please select correct folder")
            FiLeSenderWindow.update()
            return
        window1 = Toplevel(FiLeSenderWindow)
        window1.title('All PDF files in folder, please select another folder if wrong')
        window1.geometry("900x700+500+200") 
        window1.configure(bg="RoyalBlue4")
        window1.attributes('-topmost',1)
        ListofPDFfilesinFolder = Listbox(window1, bg="royalblue1",width=80, height=31, selectmode='single', font=('Times', 14))
        ListofPDFfilesinFolder.pack()
        ListofPDFfilesinFolder.insert(END,'All PDF files in selected folder, please choose another folder if wrong')
        ListofPDFfilesinFolder.insert(END,'')
        ListofPDFfilesinFolder.insert(END,'Total files in folder: '+str(TotalFiLesToBeSorted))
        ListofPDFfilesinFolder.insert(END,'')
        for x in FiLesInToSortFolder:
            ListofPDFfilesinFolder.insert(END,x)
        labetotalFiLetosend.configure(text="FiLes to be processed: "+str(TotalFiLesToBeSorted))
        SortSafety2 = False
        button_explore2.configure(bg='green4')
        FiLeSenderWindow.update()
    def Sort():
        global SortSafety1
        global SortSafety2
        if SortSafety1 == True:
            labelforERRORs.configure(text="Please select a report before sorting")
            FiLeSenderWindow.update()
            return
        if SortSafety2 == True:
            labelforERRORs.configure(text="Please select the FiLe folder before sorting")
            FiLeSenderWindow.update()
            return
        Path(FiLesFolderToSort+'\\FiLes Ready To Send\\FiLes Sent').mkdir(parents=True, exist_ok=True)
        Path(FiLesFolderToSort+'\\FiLes Ready To Send').mkdir(parents=True, exist_ok=True)
        labelforERRORs.configure(text="")
        labetotalFiLetosend.configure(text="")
        FilesRemaned = 0
        FileofErrors = 0
        for FiLes in FiLesInToSortFolder:
            RefNumber = ""
            Name1 = ""
            Name2 = ""
            NatNumber = ""
            NameInitials = ''
            page = convert_from_path(FiLesFolderToSort+"\\"+FiLes, 450,poppler_path = PopplerPath) 
            page[0].save(FiLes[:-4]+'.jpg', 'JPEG')
            x = Image.open(FiLes[:-4]+'.jpg')
            pageContent = pytesseract.image_to_string(x)
            for content in pageContent.split():
                if len(content) == 6 and content.startswith("0"): ## and is int
                    RefNumber = content
            if RefNumber == "":
                FileofErrors += 1
                x.close()
                os.remove(FiLes[:-4]+'.jpg')
                labetotalFiLetosend.configure(text='File being Processed: '+str(FilesRemaned)+'/'+str(TotalFiLesToBeSorted)+"\n Files Unable to Rename: "+str(FileofErrors)+'/'+str(TotalFiLesToBeSorted))
                os.rename(FiLesFolderToSort+"\\"+FiLes,FiLesFolderToSort+"\\"+FiLes[:-4]+" NO NUMBER FOUND"+FiLes[-4:])
                FiLeSenderWindow.update()
                continue
            Name1 = RefNumberrtoNameDict1.get(RefNumber)
            Name2 = RefNumberrtoNameDict2.get(RefNumber)
            NameInitials = RefNumberrtoNameIntialsDict.get(RefNumber)
            NatNumber =  RefNumberrtoNATDict.get(RefNumber)
            if Name1 in pageContent.split() and Name2 in pageContent.split():
                pass
            else:
                FileofErrors += 1
                x.close()
                os.remove(FiLes[:-4]+'.jpg')
                labetotalFiLetosend.configure(text='File being Processed: '+str(FilesRemaned)+'/'+str(TotalFiLesToBeSorted)+"\n Files Unable to Rename: "+str(FileofErrors)+'/'+str(TotalFiLesToBeSorted))
                os.rename(FiLesFolderToSort+"\\"+FiLes,FiLesFolderToSort+"\\"+FiLes[:-4]+" NAME NOT MATCHED, SEND MANUALLY"+FiLes[-4:])
                FiLeSenderWindow.update()
                continue
            x.close()
            os.remove(FiLes[:-4]+'.jpg')
            PDFFile = PyPDF3.PdfFileReader(FiLesFolderToSort+'/'+FiLes)
            NumberOfPages = PDFFile.numPages
            Output_PDFFile = PyPDF3.PdfFileWriter()
            for i in range(NumberOfPages):
                Output_PDFFile.addPage(PDFFile.getPage(i))
            Output_PDFFile.encrypt(NatNumber)
            Output_PDFFile.write(open(FiLesFolderToSort+'\\FiLes Ready To Send\\'+RefNumber+" "+NameInitials+".pdf", 'wb'))
            FilesRemaned += 1
            labetotalFiLetosend.configure(text='File being Processed: '+str(FilesRemaned)+'/'+str(TotalFiLesToBeSorted)+"\n Files Unable to Rename: "+str(FileofErrors)+'/'+str(TotalFiLesToBeSorted))
            os.remove(FiLesFolderToSort+"\\"+FiLes)
            FiLeSenderWindow.update()
        labetotalFiLetosend.configure(text='Total files renamed: '+str(FilesRemaned)+'/'+str(TotalFiLesToBeSorted)+"\n Files Unable to Rename: "+str(FileofErrors)+'/'+str(TotalFiLesToBeSorted))
        button_sort.configure(bg='green4')
        FiLeSenderWindow.update()
    def Send():
        global SortSafety1
        global SortSafety2
        if SortSafety1 == True or SortSafety2 == True:
            labelforERRORs.configure(text="Please select a report & folder before sending")
            FiLeSenderWindow.update()
            return
        FiLestoSend = os.listdir(FiLesFolderToSort+'\\FiLes Ready To Send')
        ListofFiLestoEmailAddress = []
        for files4 in FiLestoSend:
            if files4.endswith('.pdf') or files4.endswith('.PDF'):
                newFileName = ''
                EmailAddress = ""
                RefNumber = ''
                NameOnthePDFFile = ''
                files1 = ''
                files1 = Path(files4).stem 
                RefNumber = files1[:6]
                NameOnthePDFFile = files1[7:]
                newFileName = RefNumber+' '+NameOnthePDFFile
                EmailAddress = RefNumberrToEmailDict.get(RefNumber)
                if EmailAddress == None:
                    labelforERRORs.configure(text=str(newFileName)+" Number not found on Report")
                    FiLeSenderWindow.update()
                    return
                ListofFiLestoEmailAddress.append(newFileName+'    ---->    '+EmailAddress)
                files1 = ''
                RefNumber = ''
                NameOnthePDFFile = ''
                newFileName = ''
                EmailAddress = ""  
        window2 = Toplevel(FiLeSenderWindow)
        window2.title('List of FiLes and Address for them to be sent to')
        window2.geometry("900x700+500+200") 
        window2.configure(bg="RoyalBlue4")
        window2.attributes('-topmost',1)
        ListofFiLetoEmail = Listbox(window2, bg="royalblue1",width=80, height=25, selectmode='single', font=('Times', 14))
        ListofFiLetoEmail.pack()
        ListofFiLetoEmail.insert(END,'Total FiLes to be sent '+str(len(FiLestoSend)-1))
        ListofFiLetoEmail.insert(END,'')
        for x in ListofFiLestoEmailAddress:
            ListofFiLetoEmail.insert(END,x)
        def SendSecond():
            global SortSafety1
            FiLeSentCOunter = 0
            for FiLestos in FiLestoSend:
                if FiLestos.endswith('.pdf') or FiLestos.endswith('.PDF'):
                    RefNumber = ""
                    EmailAddress = ""
                    file = Path(FiLesFolderToSort+'\\FiLes Ready To Send\\'+FiLestos).stem           #### CHANGE FOLDER NAME HERE 
                    RefNumber = file[:6]
                    EmailAddress = RefNumberrToEmailDict.get(RefNumber)
                    outlook = win32.Dispatch('outlook.application')
                    mail = outlook.CreateItem(0)
                    mail.To = EmailAddress
                    mail.Subject = ''
                    mail.Body = ''
                    mail.Attachments.Add(FiLesFolderToSort+'\\FiLes Ready To Send\\'+FiLestos)
                    mail.CC = ""
                    mail.Send()
                    FiLeSentCOunter += 1
                    LabelSendSecondForSent.configure(text = "Amount of files sent: %d/%d" % (FiLeSentCOunter,TotalFiLesToBeSorted))
                    RefNumber = ""
                    EmailAddress = ""
                    shutil.move(FiLesFolderToSort+'\\FiLes Ready To Send\\'+FiLestos,FiLesFolderToSort+'\\FiLes Ready To Send\\FiLes Sent\\'+FiLestos)
                    FiLeSenderWindow.update()
                    time.sleep(1)
            LabelSendSecondForSent.configure(text = "All emails sent %d/%d\n\nPlease remove all files in 'FiLes To Send', ready for next process" % (FiLeSentCOunter,TotalFiLesToBeSorted))
            SortSafety1 = True
            FiLeSenderWindow.update()
        SendFiLeMain2135 = Button(window2,text = "Send",bg = 'blue',width = 30,height = 2,command = SendSecond, font=('Times', 15, 'bold'), fg = "yellow2")
        LabelSendSecondForSent = Label(window2,text = "",width = 75,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
        LabelSendSecondForSent.pack()
        SendFiLeMain2135.pack()
    label_file_explorer = Label(FiLeSenderWindow,text = "O",width = 100, height = 4,fg = "white",bg = 'RoyalBlue4', font=('Times', 13)) #RoyalBlue4
    button_explore = Button(FiLeSenderWindow,text = "Select  Report",bg = 'red3',width = 30, height = 2,command = browser, font=('Times', 15, 'bold'), fg = "yellow2")
    button_explore2 = Button(FiLeSenderWindow,text = "Select FiLe Folder",bg = 'red3',width = 30, height = 2,command = browser2, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exit = Button(FiLeSenderWindow,text = "Exit",bg = 'snow4',width = 30,height = 2,command = sys.exit, font=('Times', 15, 'bold'), fg = "black")
    button_sort = Button(FiLeSenderWindow,text = "Sort",bg = 'blue',width = 30,height = 2,command = Sort, font=('Times', 15, 'bold'), fg = "yellow2")
    button_send = Button(FiLeSenderWindow,text = "Send",bg = 'blue',width = 30,height = 2,command = Send, font=('Times', 15, 'bold'), fg = "yellow2")
    labelfileopened = Label(FiLeSenderWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labetotalFiLetosend = Label(FiLeSenderWindow,text = "",width = 50, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelforERRORs = Label(FiLeSenderWindow,text = "",width = 75, height = 2,fg = "red",bg = 'RoyalBlue4', font=('Times', 16))
    labetotalFiLetosend = Label(FiLeSenderWindow,text = "",width = 75,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    LabelSpace1FiLe = Label(FiLeSenderWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace2FiLe = Label(FiLeSenderWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace3FiLe = Label(FiLeSenderWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace4FiLe = Label(FiLeSenderWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    TitleImage = os.getcwd()+'\\1519797862804.jpg'
    img = Image.open(TitleImage)
    img = img.resize((500, 100), Image.LANCZOS)
    img = ImageTk.PhotoImage(img)
    labelimage = Label(FiLeSenderWindow, image = img,width = 500, height = 100)
    button_send.configure(bg='blue')
    def allpacks():
        labelimage.pack()
        label_file_explorer.pack()
        button_explore.pack()
        LabelSpace1FiLe.pack()
        button_explore2.pack()
        LabelSpace2FiLe.pack()
        button_sort.pack()
        LabelSpace3FiLe.pack()
        button_exit.pack()
        LabelSpace4FiLe.pack()
        button_send.pack()
        labelforERRORs.pack()
        labelfileopened.pack()
        labetotalFiLetosend.pack()
        labetotalFiLetosend.pack()
    allpacks()
    FiLeSenderWindow.mainloop()
def FILESelected():
    FILEWindow=Toplevel(MainWindow)
    FILEWindow.title("Sending emails")
    FILEWindow.geometry("900x800+500+100")
    FILEWindow.configure(bg="RoyalBlue4")
    FILEWindow.attributes('-topmost',1)
    global FILESendSafety1
    global FILESendSafety2
    FILESendSafety1 = True
    FILESendSafety2 = True
    def FILEExcelFile():
        global FILEExcelFIle
        global ExcelFileFILE
        global ExcelSheetFILE
        global AddressListFILE
        global TotalEmailAdressesFILE
        global FILESendSafety1
        button_exploreExcelFILE.configure(bg='red3')
        button_exploreDOCxFILE.configure(bg='red3')
        button_sendFILE.configure(bg='red3')
        FILESendSafety1 = True
        labelERRORsfoRFILE.configure(text='')
        AddressListFILE = []
        FILEExcelFIle = filedialog.askopenfilename(parent=FILEWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select file",filetypes = (("Excel","*xlsx"),("all files","*.*")))
        FILEExcelFIle0 = Path(FILEExcelFIle).stem 
        labelfileopnedFILE.configure(text="File Opened: "+FILEExcelFIle0)
        ExcelFileFILE = openpyxl.load_workbook(FILEExcelFIle)
        ExcelSheetFILE = ExcelFileFILE.active
        for cell in ExcelSheetFILE['N']:
            if cell.value != None and '@' in cell.value:
                AddressListFILE.append(cell.value)
        TotalEmailAdressesFILE = len(AddressListFILE)
        if TotalEmailAdressesFILE > 0:
            pass
        else:
            labelERRORsfoRFILE.configure(text='ERROR: No email address found in Col "N"')
            FILEWindow.update()
            FILESendSafety1 = True
            return
        FILEWindo1 = Toplevel(FILEWindow)
        FILEWindo1.title('All PDF files in folder, please select another folder if wrong')
        FILEWindo1.geometry("900x700+500+200") 
        FILEWindo1.configure(bg="RoyalBlue4")
        FILEWindo1.attributes('-topmost',1)
        ListofEmailAddressListBoxFILE = Listbox(FILEWindo1, bg="royalblue1",width=80, height=31, selectmode='single', font=('Times', 14))
        ListofEmailAddressListBoxFILE.pack()
        ListofEmailAddressListBoxFILE.insert(END,'All Email Addressses in selected excel file, please choose another file if wrong')
        ListofEmailAddressListBoxFILE.insert(END,'')
        ListofEmailAddressListBoxFILE.insert(END,'Total emails to send to: '+str(TotalEmailAdressesFILE))
        ListofEmailAddressListBoxFILE.insert(END,'')
        for xFILE in AddressListFILE:
            ListofEmailAddressListBoxFILE.insert(END,xFILE)
        labelsentcounterFILE.configure(text="Total Email Addresses to send to: "+str(TotalEmailAdressesFILE))
        FILESendSafety1 = False
        button_exploreExcelFILE.configure(bg='green4')
    def SelectFILEDocx():
        global FILESendSafety2
        global enrollmentletterFILE
        button_exploreDOCxFILE.configure(bg='red3')
        button_sendFILE.configure(bg='red3')
        FILESendSafety2 = True
        enrollmentletterFILE = ''
        labelERRORsfoRFILE.configure(text='')
        if FILESendSafety1 == True:
            labelERRORsfoRFILE.configure(text='Please select the excel report first')
            button_exploreDOCxFILE.configure(bg='red3')
            return
        try:
            enrollmentletterFILE = filedialog.askopenfilename(parent=FILEWindow, initialdir = "C:\\Users\\nickb\\Desktop",title = "Select file",filetypes = (("DOCx","*docx"),))
            filename0FILE = Path(enrollmentletterFILE).stem 
            labelfileopnedDocxFILE.configure(text="File To Be Sent: "+filename0FILE)
        except:
            labelERRORsfoRFILE.configure(text='Please select a valid DOCx file')
            return
        if len(enrollmentletterFILE) == 0:
            FILESendSafety2 = True
            button_exploreDOCxFILE.configure(bg='red3')
            labelERRORsfoRFILE.configure(text='Please select a valid DOCx file')
            return
        else:
            FILESendSafety2 = False
            button_exploreDOCxFILE.configure(bg='green4')
            button_sendFILE.configure(bg='blue')
    def SendFILE():
        global FILESendSafety1
        global FILESendSafety2
        if FILESendSafety1 == True:
            labelERRORsfoRFILE.configure(text='Please select Excel file before sending')
            FILEWindow.update()
            return
        if FILESendSafety2 == True:
            labelERRORsfoRFILE.configure(text='Please select DOCx before sending')
            FILEWindow.update()
            return
        else:         
            global TotalEmailAdressesFILE
            global FILElettersent
            global enrollmentletterFILE
            FILElettersent = 0
            for FILEEmails in AddressListFILE:
                outlook = win32.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)
                mail.To = FILEEmails
                mail.Subject = ''
                mail.Body = ''
                mail.CC = ""
                mail.Attachments.Add(enrollmentletterFILE)
                mail.Send()
                FILElettersent += 1
                labelsentcounterFILE.configure(text = "Amount of files sent: %d/%d" % (FILElettersent,TotalEmailAdressesFILE))
                time.sleep(1)
                FILEWindow.update()
            labelsentcounterFILE.configure(text = "All emails sent %d/%d" % (FILElettersent,TotalEmailAdressesFILE))
            FILESendSafety1 = 'On'
            FILESendSafety2 = 'On'
            button_exploreExcelFILE.configure(bg='red3')
            button_exploreDOCxFILE.configure(bg='red3')
            button_sendFILE.configure(bg='red3')
            FILEWindow.update()
    label_file_explorerFILE = Label(FILEWindow,text = "",width = 100, height = 4,fg = "white",bg = 'RoyalBlue4', font=('Times', 13))
    button_exploreExcelFILE = Button(FILEWindow,text = "Select Excel File for Sending",bg = 'red3',width = 30, height = 2,command = FILEExcelFile, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exploreDOCxFILE = Button(FILEWindow,text = "Select FILE template file to be Sent",bg = 'red3',width = 30, height = 2,command = SelectFILEDocx, font=('Times', 15, 'bold'), fg = "yellow2")
    button_exitFILE = Button(FILEWindow,text = "Exit",bg = 'snow4',width = 30,height = 2,command = sys.exit, font=('Times', 15, 'bold'), fg = "black")
    button_sendFILE = Button(FILEWindow,text = "Send",bg = 'red3',width = 30,height = 2,command = SendFILE, font=('Times', 15, 'bold'), fg = "yellow2")
    labelERRORsfoRFILE = Label(FILEWindow,text = "",width = 75, height = 2,fg = "red",bg = 'RoyalBlue4', font=('Times', 16))
    labelfileopnedFILE = Label(FILEWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelsentcounterFILE = Label(FILEWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    labelfileopnedDocxFILE = Label(FILEWindow,text = "",width = 75, height = 2,fg = "white",bg = 'RoyalBlue4', font=('Times', 16))
    LabelSpace1FILE = Label(FILEWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace2FILE = Label(FILEWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    LabelSpace3FILE = Label(FILEWindow,text = "",width = 75, height = 1,fg = "red",bg = 'RoyalBlue4')
    TitleImageFILE = os.getcwd()+'\\1519797862804.jpg'
    imgFILE = Image.open(TitleImageFILE)
    imgFILE = imgFILE.resize((500, 100), Image.LANCZOS)
    imgFILE = ImageTk.PhotoImage(imgFILE)
    labelimageFILE = Label(FILEWindow, image = imgFILE,width = 500, height = 100)
    def allpacks():
        labelimageFILE.pack()
        label_file_explorerFILE.pack()
        button_exploreExcelFILE.pack()
        LabelSpace1FILE.pack()
        button_exploreDOCxFILE.pack()
        LabelSpace2FILE.pack()
        button_exitFILE.pack()
        LabelSpace3FILE.pack()
        button_sendFILE.pack()
        labelERRORsfoRFILE.pack()
        labelfileopnedFILE.pack()
        labelfileopnedDocxFILE.pack()
        labelsentcounterFILE.pack()
    allpacks()
    FILEWindow.mainloop()   
TitleImage = os.getcwd()+'\\1519797862804.jpg'  
img = Image.open(TitleImage)
img = img.resize((500, 100), Image.LANCZOS)
img = ImageTk.PhotoImage(img)
labelimage = Label(MainWindow, image = img,width = 500, height = 100)
Labelfill = Label(MainWindow, width = 30, height = 3,bg = 'RoyalBlue4')
Labelfill1 = Label(MainWindow, width = 30, height = 3,bg = 'RoyalBlue4')
FiLeWindowButton = Button(MainWindow,text = "Send FiLes",bg = 'RoyalBlue1',width = 30,height = 2, font=('Times', 15, 'bold'), fg = "yellow2",command = FiLeSelected) #command = Sort
FILEWindowButton = Button(MainWindow,text = "Send FILE files",bg = 'RoyalBlue1',width = 30,height = 2, font=('Times', 15, 'bold'), fg = "yellow2",command = FILESelected) #command = FiLeSelected
labelimage.pack()
Labelfill.pack()
FiLeWindowButton.pack()
Labelfill1.pack()
FILEWindowButton.pack()
MainWindow.mainloop()

